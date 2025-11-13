#!/usr/bin/env python3

"""
SuperAudit - Scale Computing HyperCore VM Inventory Tool

Collects a full inventory of virtual machines across a cluster, including:
  - VM metadata (name, UUID, description, status, CPU, memory)
  - Disk information (drive type, mount point, size, SSD tiering, snapshot size)
  - Network configuration (IP addresses, VLANs, mounted ISOs)
  - Snapshot schedules and replication partners
  - Node CPU usage and current VM placement

Output is saved as an Excel workbook (.xlsx) with multiple formatted worksheets:
  - Summary Dashboard (statistics, warnings, collection info)
  - VM Inventory (complete VM audit with conditional formatting)
  - Node Hardware (node specifications and health)
  - Drive Health (disk health and temperature monitoring)
  - ISO Library (ISO inventory and mount status)

Uses the Scale Computing REST API instead of CLI commands.

USAGE:
    # Interactive mode (prompts for credentials) - MOST SECURE - Generates Excel by default
    ./SuperAudit_API.py

    # Using environment variables - SECURE for automation
    export SCALE_USER=admin
    export SCALE_PASSWORD=yourpassword
    ./SuperAudit_API.py -n 10.205.109.101

    # Using .netrc file - SECURE standard method
    # Create ~/.netrc with permissions 0600:
    # machine 10.205.109.101
    # login admin
    # password yourpassword
    ./SuperAudit_API.py -n 10.205.109.101

    # Legacy CSV output mode
    ./SuperAudit_API.py -n 10.205.109.101 -u admin --format csv

OPTIONS:
    -n, --node HOST        Cluster node hostname or IP
    -u, --user USERNAME    Username for authentication
    -p, --password PASS    Password for authentication
    -q, --quiet            Run in quiet mode (less verbose output)
    --format {xlsx,csv}    Output format: xlsx (Excel) or csv (default: xlsx)
    --no-verify-ssl        Disable SSL certificate verification
    --ca-bundle PATH       Path to CA certificate bundle for self-signed certificates
    --include-nodes        [Deprecated] Generate node hardware report (auto-included in Excel)
    --include-drives       [Deprecated] Generate drive health report (auto-included in Excel)
    --include-isos         [Deprecated] Generate ISO library report (auto-included in Excel)
    --all-reports          [Deprecated] Generate all reports (default behavior in Excel)
    --summary-only         Show cluster and audit summary only (no file generation)
    --warnings             Show only warnings (exit 0 if none, 1 if found)
    -v, --version          Show version and exit
    -h, --help             Show help message and exit

EXAMPLES:
    # Generate Excel workbook with all sheets (DEFAULT)
    ./SuperAudit_API.py -n 10.205.109.101 -u admin
    # Output: superaudit_clustername.xlsx

    # Generate legacy CSV output
    ./SuperAudit_API.py -n cluster.local -u admin --format csv
    # Output: superaudit_clustername.csv

    # Using environment variables (secure for scripts)
    export SCALE_USER=admin SCALE_PASSWORD=secret
    ./SuperAudit_API.py -n cluster.local
    # Output: Excel workbook with summary dashboard and all sheets

    # Using .netrc file (secure, standard Unix method)
    ./SuperAudit_API.py -n cluster.local

    # Using custom CA bundle for self-signed certificates (SECURE)
    ./SuperAudit_API.py -n 10.205.109.101 -u admin --ca-bundle /path/to/scale_cert.pem

    # Quick health check (summary only, no files)
    ./SuperAudit_API.py -n cluster.local -u admin -p password --summary-only

    # Check for warnings only (useful for monitoring scripts)
    ./SuperAudit_API.py -n cluster.local -u admin -p password --warnings

SECURITY:
    Credential security (in order of preference):

    1. Interactive prompts - Password never visible (MOST SECURE)
       ./SuperAudit_API.py -n cluster.local -u admin

    2. .netrc file - Standard Unix credentials file (SECURE)
       Create ~/.netrc with mode 0600:
       machine cluster.local
       login admin
       password yourpassword

    3. Environment variables - Good for automation (SECURE)
       export SCALE_USER=admin
       export SCALE_PASSWORD=yourpassword

    4. Command-line arguments - AVOID IN PRODUCTION (INSECURE)
       -p password exposes credentials in:
       - Process lists (ps, top, htop)
       - Shell history
       - System logs
       Use only for testing!

    SSL/TLS Certificate Verification (self-signed certificates):

    For self-signed certificates, use --ca-bundle (RECOMMENDED):
       # Download your server's certificate
       echo | openssl s_client -connect 10.205.109.101:443 -showcerts 2>/dev/null | \
         openssl x509 -outform PEM > scale_cert.pem

       # Use the certificate bundle
       ./SuperAudit_API.py -n 10.205.109.101 -u admin --ca-bundle scale_cert.pem

    Benefits of --ca-bundle:
       ✓ Credentials remain encrypted via TLS
       ✓ Verifies you're connecting to the correct server
       ✓ Protects against man-in-the-middle attacks

    Alternative (NOT RECOMMENDED):
       --no-verify-ssl disables certificate verification
       - Credentials are still encrypted
       - But vulnerable to man-in-the-middle attacks
       - Only use for testing!

REQUIREMENTS:
    - Python 3.6 or later
    - openpyxl library for Excel output (pip install openpyxl)
    - Access to Scale Computing HyperCore cluster REST API
    - Valid cluster credentials

EXCEL FEATURES (v7.0+):
    - Multi-sheet workbook with 6 worksheets:
      1. Summary Dashboard - Key statistics, warnings, system utilization
      2. VM Inventory - Complete VM audit with 30 data fields
      3. Node Hardware - Node specifications and health
      4. Drive Health - Disk health and temperature monitoring
      5. ISO Library - ISO inventory and mount status
      6. Warnings & Recommendations - Categorized issues with solutions
    - Conditional formatting (color-coded by severity and status)
    - Freeze panes and auto-filter on all data sheets
    - Auto-sized columns for readability
    - Professional formatting with colored headers

FILTERING & EXPORT (v7.3+):
    - Filter VMs by: state, type, tags, name, node
    - JSON export for automation and dashboards
    - Summary export for monitoring integrations
    - Combine filters for targeted audits

TESTING & VALIDATION (v7.4+):
    - Test connection before running full audit (--test-connection)
    - Dry run mode to validate filters (--dry-run)
    - Enhanced error messages with troubleshooting hints

CHANGELOG:
    v7.4 - Test connection and dry run modes, enhanced UX
    v7.3 - VM filtering, JSON export, summary statistics export
    v7.2 - Compliance reporting, capacity planning, categorized warnings
    v7.1 - Enhanced data collection (boot order, HA, network details, cache mode)
    v7.0 - Excel output with multi-sheet workbooks, bug fixes, session handling
    v6.1 - Bug fixes and improvements
    v6.0 - Initial release with REST API support

Author: Scale Computing Support Team
Version: 7.4
"""

import argparse
import base64
import csv
import datetime
import http.client as http
import json
import os
import ssl
import sys
import time
from getpass import getpass
from pathlib import Path
from typing import Dict, List, Optional, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ======== GLOBAL VARIABLES ========
VERSION = "8.0"
DEFAULT_OUTPUT_FILE = None  # Will be set based on cluster name

# ANSI color codes
COLOR_RESET = "\033[0m"
COLOR_BOLD = "\033[1m"
COLOR_YELLOW = "\033[33m"
COLOR_BLUE = "\033[34m"
COLOR_GREEN = "\033[32m"
COLOR_CYAN = "\033[36m"
COLOR_RED = "\033[31m"

BOX_WIDTH = 70


# ======== EXCEPTION CLASSES ========

class InternalException(Exception):
    """Base exception for internal errors"""
    pass


class HTTPResponseException(InternalException):
    """Exception for HTTP response errors"""
    def __init__(self, response, body=None):
        self.response = response
        self.status = response.status
        self.body = body if body else ""

    def __str__(self):
        return f"HTTP {self.status}: {self.body}"


class TaskException(InternalException):
    """Exception for task errors"""
    def __init__(self, tag, message, parameters):
        self.tag = tag
        self.message = message
        self.parameters = parameters

    def __str__(self):
        return f'{self.tag} "{self.message}" {self.parameters}'


# ======== API CLIENT CLASS ========

class ScaleAPIClient:
    """Scale Computing REST API Client"""

    def __init__(self, host: str, username: str, password: str, verify_ssl: bool = False, ca_bundle: Optional[str] = None):
        self.host = host
        self.username = username
        self.password = password
        self.verify_ssl = verify_ssl
        self.ca_bundle = ca_bundle
        self.session_cookie = None
        self.base_url = f"/rest/v1"

        # Create SSL context
        self.ssl_context = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)

        if ca_bundle:
            # Load custom CA bundle (SECURE - verifies with your self-signed cert)
            self.ssl_context.load_verify_locations(cafile=ca_bundle)
        elif not verify_ssl:
            # Disable verification (INSECURE - but credentials still encrypted)
            self.ssl_context.check_hostname = False
            self.ssl_context.verify_mode = ssl.CERT_NONE
        # else: use default system CA bundle

        # Connection timeout
        self.timeout = 120

    def _get_connection(self) -> http.HTTPSConnection:
        """Create a new HTTPS connection"""
        return http.HTTPSConnection(
            self.host,
            timeout=self.timeout,
            context=self.ssl_context
        )

    def _get_auth_header(self) -> str:
        """Generate Basic Auth header"""
        credentials = f"{self.username}:{self.password}"
        encoded = base64.b64encode(credentials.encode('utf-8')).decode('utf-8')
        return f"Basic {encoded}"

    def login(self) -> bool:
        """Authenticate and get session cookie"""
        connection = self._get_connection()

        auth_header = self._get_auth_header()
        payload = json.dumps({
            "username": self.username,
            "password": self.password,
            "useOIDC": False
        })

        headers = {
            'Authorization': auth_header,
            'Content-Type': 'application/json'
        }

        try:
            connection.request('POST', f'{self.base_url}/login', payload, headers)
            response = connection.getresponse()
            body = response.read().decode('utf-8')

            if response.status != 200:
                raise HTTPResponseException(response, body)

            # Extract session cookie
            cookies = response.getheader('Set-Cookie')
            if cookies and 'sessionID=' in cookies:
                # Parse sessionID from cookie string
                for cookie in cookies.split(';'):
                    if 'sessionID=' in cookie:
                        self.session_cookie = cookie.split('sessionID=')[1].split(';')[0]
                        break

            return True

        except Exception as e:
            raise Exception(f"Login failed: {e}")
        finally:
            connection.close()

    def logout(self):
        """Logout and destroy session"""
        if not self.session_cookie:
            return

        connection = self._get_connection()
        headers = {
            'Cookie': f'sessionID={self.session_cookie}'
        }

        try:
            connection.request('POST', f'{self.base_url}/logout', None, headers)
            response = connection.getresponse()
            response.read()  # Consume response
            self.session_cookie = None
        except:
            pass  # Ignore logout errors
        finally:
            connection.close()

    def get(self, endpoint: str, retry_auth: bool = True) -> Any:
        """Perform GET request with automatic session timeout handling"""
        connection = self._get_connection()

        headers = {
            'Content-Type': 'application/json'
        }

        if self.session_cookie:
            headers['Cookie'] = f'sessionID={self.session_cookie}'
        else:
            headers['Authorization'] = self._get_auth_header()

        try:
            url = f"{self.base_url}/{endpoint}" if not endpoint.startswith('/') else endpoint
            connection.request('GET', url, None, headers)
            response = connection.getresponse()
            body = response.read().decode('utf-8')

            # Handle session timeout - retry once with re-authentication
            if response.status == 401 and retry_auth and self.session_cookie:
                connection.close()
                # Session expired, re-authenticate and retry
                self.login()
                return self.get(endpoint, retry_auth=False)  # Retry once without further retries

            if response.status != 200:
                raise HTTPResponseException(response, body)

            return json.loads(body) if body else None

        finally:
            connection.close()

    def post(self, endpoint: str, data: Dict, retry_auth: bool = True) -> Any:
        """Perform POST request with automatic session timeout handling"""
        connection = self._get_connection()

        headers = {
            'Content-Type': 'application/json'
        }

        if self.session_cookie:
            headers['Cookie'] = f'sessionID={self.session_cookie}'
        else:
            headers['Authorization'] = self._get_auth_header()

        try:
            url = f"{self.base_url}/{endpoint}" if not endpoint.startswith('/') else endpoint
            payload = json.dumps(data)
            connection.request('POST', url, payload, headers)
            response = connection.getresponse()
            body = response.read().decode('utf-8')

            # Handle session timeout - retry once with re-authentication
            if response.status == 401 and retry_auth and self.session_cookie:
                connection.close()
                # Session expired, re-authenticate and retry
                self.login()
                return self.post(endpoint, data, retry_auth=False)  # Retry once without further retries

            if response.status not in [200, 201]:
                raise HTTPResponseException(response, body)

            return json.loads(body) if body else None

        finally:
            connection.close()


# ======== HELPER FUNCTIONS ========

def draw_line(width: int = BOX_WIDTH):
    """Draw a horizontal line"""
    print("-" * width)


def draw_centered_text(text: str, width: int = BOX_WIDTH):
    """Draw centered text"""
    padding = (width - len(text)) // 2
    print(" " * padding + text)


def log_info(message: str, quiet: bool = False):
    """Log informational message"""
    if not quiet:
        print(message)


def sanitize_filename(name: str) -> str:
    """Sanitize cluster name for use in filename"""
    import re
    # Replace spaces and special characters with underscores
    sanitized = re.sub(r'[^\w\-.]', '_', name)
    # Remove consecutive underscores
    sanitized = re.sub(r'_+', '_', sanitized)
    # Remove leading/trailing underscores
    sanitized = sanitized.strip('_')
    # Convert to lowercase for consistency
    sanitized = sanitized.lower()
    # If empty after sanitization, use 'cluster'
    return sanitized if sanitized else 'cluster'


def convert_to_gb(size_bytes: int) -> float:
    """Convert bytes to GB (base 1000)"""
    if not size_bytes:
        return 0.0
    return round(size_bytes / 1_000_000_000, 2)


def convert_memory_to_gb(mem_bytes: int) -> str:
    """Convert memory bytes to GB with appropriate decimal places"""
    if not mem_bytes:
        return "0"

    gb_value = mem_bytes / 1_073_741_824  # 1024^3 for GiB

    # Convert to decimal GB (base 1000) like the bash script
    gb_value = mem_bytes / 1_000_000_000

    # Strip .00 for whole numbers
    if gb_value == int(gb_value):
        return str(int(gb_value))
    else:
        return f"{gb_value:.3f}".rstrip('0').rstrip('.')


def map_ssd_priority(tier: int) -> str:
    """Map SSD tiering value to display value"""
    ssd_map = {
        0: "0 - OFF (0)",
        1: "1 - MINIMAL (1)",
        2: "2 - VERY LOW (2)",
        4: "3 - LOW (4)",
        8: "4 - NORMAL (8)",
        16: "5 - HIGH (16)",
        32: "6 - VERY HIGH (32)",
        64: "7 - EXTREME (64)",
        128: "8 - ABSURD (128)",
        256: "9 - HYPERSPEED (256)",
        1024: "10 - LUDICROUS SPEED (1024)",
        10240: "11 - THESE GO TO 11 (10240)"
    }
    return ssd_map.get(tier, str(tier))


def clean_ipv6(ip: str) -> str:
    """Clean IPv6 addresses by removing %vlan portion"""
    if '%' in ip:
        return ip.split('%')[0]
    return ip


def vm_matches_filter(vm_name: str, vm_state: str, vm_type: str, vm_tags: str, node_lan_ip: str, args) -> bool:
    """Check if VM matches filter criteria (Push 4)"""
    # Filter by state
    if args.filter_state and vm_state != args.filter_state:
        return False

    # Filter by type
    if args.filter_type and vm_type != args.filter_type:
        return False

    # Filter by tag (partial match)
    if args.filter_tag and args.filter_tag.lower() not in vm_tags.lower():
        return False

    # Filter by name (partial match, case-insensitive)
    if args.filter_name and args.filter_name.lower() not in vm_name.lower():
        return False

    # Filter by node
    if args.filter_node and node_lan_ip != args.filter_node:
        return False

    return True


def get_credentials_from_netrc(host: str) -> Optional[tuple]:
    """
    Try to get credentials from .netrc file

    Format in ~/.netrc:
        machine cluster.domain.com
        login admin
        password yourpassword

    File permissions must be 0600 for security
    """
    netrc_path = Path.home() / '.netrc'

    if not netrc_path.exists():
        return None

    # Check file permissions (must be 0600 for security)
    stat_info = netrc_path.stat()
    if stat_info.st_mode & 0o077:  # Check if group/other have any permissions
        print(f"{COLOR_YELLOW}Warning: .netrc file has insecure permissions. Should be 0600.{COLOR_RESET}")
        print(f"Fix with: chmod 600 {netrc_path}")
        return None

    try:
        import netrc
        authenticators = netrc.netrc(str(netrc_path))
        auth = authenticators.authenticators(host)
        if auth:
            return (auth[0], auth[2])  # (login, password)
    except Exception as e:
        # Silently fail - .netrc is optional
        pass

    return None


def get_credentials_from_env() -> Optional[tuple]:
    """
    Get credentials from environment variables

    Environment variables:
        SCALE_USER or HYPERCORE_USER
        SCALE_PASSWORD or HYPERCORE_PASSWORD
    """
    username = os.environ.get('SCALE_USER') or os.environ.get('HYPERCORE_USER')
    password = os.environ.get('SCALE_PASSWORD') or os.environ.get('HYPERCORE_PASSWORD')

    if username and password:
        return (username, password)

    return None


def show_progress(current: int, total: int, vm_name: str = "", quiet: bool = False, stats: 'AuditStatistics' = None):
    """Show progress bar with ETA and speed"""
    if quiet:
        return

    width = BOX_WIDTH - 20
    percent = int(current * 100 / total) if total > 0 else 0
    completed = int(width * current / total) if total > 0 else 0

    bar = "#" * completed
    bar = bar.ljust(width)

    # Calculate ETA and speed
    eta_str = ""
    speed_str = ""
    if stats and current > 0:
        vms_per_sec = stats.get_vms_per_second()
        if vms_per_sec > 0:
            speed_str = f" {COLOR_GREEN}[{vms_per_sec:.1f} VMs/s]{COLOR_RESET}"
            eta_seconds = stats.get_eta_seconds(current, total)
            if eta_seconds > 0:
                eta_str = f" {COLOR_YELLOW}ETA: {format_time(eta_seconds)}{COLOR_RESET}"

    vm_display = f" {COLOR_CYAN}{vm_name[:30]}{COLOR_RESET}" if vm_name else ""

    print(f"\r{COLOR_BLUE}[{bar}]{COLOR_RESET} {COLOR_YELLOW}{percent:3d}%{COLOR_RESET} ({current}/{total}){speed_str}{eta_str}{vm_display}", end='', flush=True)


# ======== DATA COLLECTION FUNCTIONS ========

def get_cluster_info(client: ScaleAPIClient) -> Dict[str, str]:
    """Get cluster information"""
    info = {'name': 'Unknown', 'company': 'Unknown', 'icosVersion': 'Unknown'}

    try:
        # Try to get cluster registration data
        cluster_data = client.get('Cluster')
        if cluster_data and len(cluster_data) > 0:
            cluster = cluster_data[0]
            info['name'] = cluster.get('clusterName', 'Unknown')
            info['company'] = cluster.get('companyName', 'Unknown')
            info['icosVersion'] = cluster.get('icosVersion', 'Unknown')
    except:
        pass

    # If we didn't get icosVersion from Cluster, try to get it from nodes
    if info['icosVersion'] == 'Unknown':
        try:
            nodes = client.get('Node')
            if nodes and len(nodes) > 0:
                info['icosVersion'] = nodes[0].get('icosVersion', 'Unknown')
        except:
            pass

    return info


def get_nodes(client: ScaleAPIClient) -> List[Dict]:
    """Get all node information"""
    nodes = client.get('Node')
    return nodes if nodes else []


def get_drives(client: ScaleAPIClient) -> List[Dict]:
    """Get all drive information"""
    drives = client.get('Drive')
    return drives if drives else []


def get_conditions(client: ScaleAPIClient) -> List[Dict]:
    """Get all active conditions (filtering done on frontend)"""
    # Use filter to only get conditions where value=true (currently active)
    conditions = client.get('Condition/filter?includeSet=true')
    return conditions if conditions else []


def get_vms(client: ScaleAPIClient) -> List[Dict]:
    """Get all VM information"""
    vms = client.get('VirDomain')
    return vms if vms else []




def get_replication_connections(client: ScaleAPIClient) -> Dict[str, str]:
    """Get replication connections, indexed by UUID"""
    try:
        # Fixed: Use correct endpoint 'RemoteClusterConnection' instead of 'VMReplicationConnection'
        connections = client.get('RemoteClusterConnection')
        if not connections:
            return {}

        conn_map = {}
        for conn in connections:
            # Map connection UUID to remote cluster name
            uuid = conn.get('uuid', '')
            name = conn.get('remoteClusterName', '')
            if uuid:
                conn_map[uuid] = name

        return conn_map
    except:
        return {}


def get_snapshot_schedules(client: ScaleAPIClient) -> Dict[str, str]:
    """Get snapshot schedules, indexed by UUID"""
    try:
        schedules = client.get('VirDomainSnapshotSchedule')
        if not schedules:
            return {}

        schedule_map = {}
        for schedule in schedules:
            schedule_map[schedule.get('uuid', '')] = schedule.get('name', '')

        return schedule_map
    except:
        return {}


# ======== STATISTICS TRACKING ========

class AuditStatistics:
    """Track statistics during audit with system utilization metrics"""
    def __init__(self):
        self.total_vms = 0
        self.running_vms = 0
        self.stopped_vms = 0
        self.total_memory_gb = 0
        self.total_storage_capacity_gb = 0
        self.total_storage_used_gb = 0
        self.vms_with_snapshots = 0
        self.vms_without_snapshots = 0
        self.start_time = time.time()

        # Categorized warnings
        self.warnings = []  # Keep for backward compatibility
        self.categorized_warnings = {
            'CRITICAL': [],  # Critical issues requiring immediate attention
            'WARNING': [],   # Important issues to address
            'INFO': []       # Informational notices
        }

        # VM type tracking
        self.vm_types = {
            'VM': 0,
            'VM REPLICA': 0,
            'TEMPLATE': 0,
            'PRODUCTION VM': 0,
            'DEV VM': 0,
            'TEST VM': 0
        }

        # System utilization metrics
        self.total_nodes = 0
        self.node_cpu_usage = []  # List of CPU usage percentages per node
        self.node_memory_usage = []  # List of memory usage percentages per node
        self.avg_cpu_usage = 0.0
        self.avg_memory_usage = 0.0
        self.max_cpu_usage = 0.0
        self.max_memory_usage = 0.0

        # Drive health metrics
        self.total_drives = 0
        self.healthy_drives = 0
        self.unhealthy_drives = 0
        self.nodes_online = 0
        self.nodes_offline = 0

    def add_vm(self, vm_state: str, mem_gb: float):
        """Add VM to statistics"""
        self.total_vms += 1
        self.total_memory_gb += mem_gb
        if vm_state == 'RUNNING':
            self.running_vms += 1
        else:
            self.stopped_vms += 1

    def add_vm_type(self, vm_type: str):
        """Track VM type"""
        if vm_type in self.vm_types:
            self.vm_types[vm_type] += 1

    def add_disk(self, capacity_gb: float, used_gb: float):
        """Add disk to statistics"""
        self.total_storage_capacity_gb += capacity_gb
        self.total_storage_used_gb += used_gb

    def add_snapshot_status(self, has_snapshots: bool):
        """Track snapshot status"""
        if has_snapshots:
            self.vms_with_snapshots += 1
        else:
            self.vms_without_snapshots += 1

    def add_warning(self, warning: str, severity: str = 'WARNING'):
        """Add a categorized warning (Push 3 enhancement)"""
        # Add to legacy warnings list for backward compatibility
        if warning not in self.warnings:
            self.warnings.append(warning)

        # Add to categorized warnings
        if severity in self.categorized_warnings:
            if warning not in self.categorized_warnings[severity]:
                self.categorized_warnings[severity].append(warning)

    def add_node_utilization(self, cpu_usage: float, memory_usage: float, network_status: str = 'ONLINE', drives: list = None):
        """Track node utilization metrics and storage from drives"""
        self.total_nodes += 1
        self.node_cpu_usage.append(cpu_usage)
        self.node_memory_usage.append(memory_usage)

        # Track node network status
        if network_status == 'ONLINE':
            self.nodes_online += 1
        else:
            self.nodes_offline += 1

        # Calculate storage from node drives (only drives that are IN)
        if drives:
            for drive in drives:
                if drive.get('currentDisposition') == 'IN':
                    capacity_bytes = drive.get('capacityBytes', 0)
                    used_bytes = drive.get('usedBytes', 0)
                    self.total_storage_capacity_gb += capacity_bytes / (1024**3)
                    self.total_storage_used_gb += used_bytes / (1024**3)

        # Update averages
        self.avg_cpu_usage = sum(self.node_cpu_usage) / len(self.node_cpu_usage)
        self.avg_memory_usage = sum(self.node_memory_usage) / len(self.node_memory_usage)

        # Update max values
        self.max_cpu_usage = max(self.node_cpu_usage)
        self.max_memory_usage = max(self.node_memory_usage)

    def add_drive_health(self, is_healthy: bool):
        """Track drive health status"""
        self.total_drives += 1
        if is_healthy:
            self.healthy_drives += 1
        else:
            self.unhealthy_drives += 1

    def get_storage_usage_percent(self) -> float:
        """Calculate storage usage percentage"""
        if self.total_storage_capacity_gb > 0:
            return (self.total_storage_used_gb / self.total_storage_capacity_gb) * 100
        return 0

    def get_storage_free_gb(self) -> float:
        """Calculate free storage in GB"""
        return self.total_storage_capacity_gb - self.total_storage_used_gb

    def get_storage_free_percent(self) -> float:
        """Calculate free storage percentage"""
        return 100 - self.get_storage_usage_percent()

    def get_elapsed_time(self) -> float:
        """Get elapsed time in seconds"""
        return time.time() - self.start_time

    def get_vms_per_second(self) -> float:
        """Calculate VMs processed per second"""
        elapsed = self.get_elapsed_time()
        if elapsed > 0:
            return self.total_vms / elapsed
        return 0

    def get_eta_seconds(self, current: int, total: int) -> float:
        """Calculate estimated time remaining"""
        vms_per_sec = self.get_vms_per_second()
        if vms_per_sec > 0:
            remaining = total - current
            return remaining / vms_per_sec
        return 0

    def get_all_warnings_count(self) -> int:
        """Get total count of all categorized warnings"""
        return sum(len(warnings) for warnings in self.categorized_warnings.values())



def format_bytes(bytes_val: int) -> str:
    """Format bytes to human readable"""
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if bytes_val < 1024.0:
            return f"{bytes_val:.1f} {unit}"
        bytes_val /= 1024.0
    return f"{bytes_val:.1f} PB"


def format_time(seconds: float) -> str:
    """Format seconds to human readable time"""
    if seconds < 60:
        return f"{int(seconds)}s"
    elif seconds < 3600:
        mins = int(seconds / 60)
        secs = int(seconds % 60)
        return f"{mins}m {secs}s"
    else:
        hours = int(seconds / 3600)
        mins = int((seconds % 3600) / 60)
        return f"{hours}h {mins}m"


def display_cluster_summary(cluster_info: Dict, nodes: List[Dict], quiet: bool = False):
    """Display cluster summary dashboard"""
    if quiet:
        return

    # Calculate total storage across all nodes
    total_capacity = sum(node.get('capacity', 0) for node in nodes)
    total_used = 0
    total_drives = 0
    ssd_drives = 0
    hdd_drives = 0

    for node in nodes:
        for drive in node.get('drives', []):
            total_drives += 1
            if drive.get('currentDisposition') == 'IN':
                total_used += drive.get('usedBytes', 0)
            if drive.get('type') == 'SSD':
                ssd_drives += 1
            else:
                hdd_drives += 1

    storage_percent = (total_used / total_capacity * 100) if total_capacity > 0 else 0

    print()
    print(f"  {COLOR_BOLD}Cluster Summary{COLOR_RESET}")
    draw_line()
    print(f"  ICOS Version:    {COLOR_YELLOW}{cluster_info.get('icosVersion', 'Unknown')}{COLOR_RESET}")
    print(f"  Nodes:           {COLOR_YELLOW}{len(nodes)}{COLOR_RESET} detected")
    print(f"  Total Drives:    {COLOR_YELLOW}{total_drives}{COLOR_RESET} ({ssd_drives} SSD, {hdd_drives} HDD)")
    print(f"  Storage:         {COLOR_YELLOW}{format_bytes(total_used)}{COLOR_RESET} / {COLOR_YELLOW}{format_bytes(total_capacity)}{COLOR_RESET} ({storage_percent:.1f}% used)")
    draw_line()
    print()


def display_summary_report(stats: AuditStatistics, output_file: str, quiet: bool = False):
    """Display final summary report with warnings"""
    if quiet:
        print(f"\nAudit completed: {stats.total_vms} VMs processed in {format_time(stats.get_elapsed_time())} - Output: {output_file}")
        if stats.warnings:
            print(f"Warnings: {len(stats.warnings)}")
        return

    # Box width (content area between borders)
    box_width = 67

    def format_line(content: str) -> str:
        """Format a line with proper padding to fit in box"""
        # Strip ANSI codes to get actual length
        import re
        clean = re.sub(r'\x1b\[[0-9;]*m', '', content)
        padding_needed = box_width - len(clean)
        if padding_needed > 0:
            content += ' ' * padding_needed
        return f"  {COLOR_BOLD}║{COLOR_RESET}{content}{COLOR_BOLD}║{COLOR_RESET}"

    print()
    draw_line()
    print()
    print(f"  {COLOR_BOLD}╔═══════════════════════════════════════════════════════════════════╗{COLOR_RESET}")
    print(f"  {COLOR_BOLD}║                      AUDIT SUMMARY                                ║{COLOR_RESET}")
    print(f"  {COLOR_BOLD}╠═══════════════════════════════════════════════════════════════════╣{COLOR_RESET}")

    # VM Statistics
    running_pct = (stats.running_vms / stats.total_vms * 100) if stats.total_vms > 0 else 0
    stopped_pct = (stats.stopped_vms / stats.total_vms * 100) if stats.total_vms > 0 else 0
    print(format_line(f" VMs Processed:      {COLOR_YELLOW}{stats.total_vms:>5}{COLOR_RESET}"))
    print(format_line(f"   Running:          {COLOR_GREEN}{stats.running_vms:>5}{COLOR_RESET} ({running_pct:>5.1f}%)"))
    print(format_line(f"   Stopped:          {COLOR_RED}{stats.stopped_vms:>5}{COLOR_RESET} ({stopped_pct:>5.1f}%)"))
    print(format_line(" "))

    # Storage Statistics
    storage_pct = stats.get_storage_usage_percent()
    storage_color = COLOR_GREEN if storage_pct < 70 else COLOR_YELLOW if storage_pct < 85 else COLOR_RED
    storage_cap_tb = stats.total_storage_capacity_gb / 1000
    storage_used_tb = stats.total_storage_used_gb / 1000
    print(format_line(f" Storage Allocated:  {COLOR_YELLOW}{storage_cap_tb:>7.1f} TB{COLOR_RESET}"))
    print(format_line(f" Storage Used:       {storage_color}{storage_used_tb:>7.1f} TB{COLOR_RESET} ({storage_color}{storage_pct:>5.1f}%{COLOR_RESET})"))
    print(format_line(" "))

    # Memory Statistics
    print(format_line(f" Memory Allocated:   {COLOR_YELLOW}{stats.total_memory_gb:>7.1f} GB{COLOR_RESET}"))
    print(format_line(" "))

    # Snapshot Statistics
    snap_pct = (stats.vms_with_snapshots / stats.total_vms * 100) if stats.total_vms > 0 else 0
    snap_color = COLOR_GREEN if snap_pct > 80 else COLOR_YELLOW if snap_pct > 50 else COLOR_RED
    print(format_line(f" VMs with Snapshots: {snap_color}{stats.vms_with_snapshots:>5}{COLOR_RESET} ({snap_pct:>5.1f}%)"))
    print(format_line(" "))

    # Timing
    elapsed = stats.get_elapsed_time()
    speed = stats.get_vms_per_second()
    time_str = format_time(elapsed)
    print(format_line(f" Execution Time:     {COLOR_CYAN}{time_str:>10}{COLOR_RESET}"))
    print(format_line(f" Processing Speed:   {COLOR_CYAN}{speed:>7.1f} VMs/s{COLOR_RESET}"))
    print(format_line(" "))

    # Output file
    output_display = output_file[:44] if len(output_file) > 44 else output_file
    print(format_line(f" Output File:        {COLOR_YELLOW}{output_display}{COLOR_RESET}"))

    # Warnings
    if stats.warnings:
        print(format_line(" "))
        print(format_line(f" {COLOR_RED}⚠  WARNINGS:{COLOR_RESET}"))
        for warning in stats.warnings[:5]:  # Show first 5 warnings
            warning_text = warning[:57]  # Leave room for bullet and spacing
            print(format_line(f"   {COLOR_RED}•{COLOR_RESET} {warning_text}"))
        if len(stats.warnings) > 5:
            remaining = len(stats.warnings) - 5
            print(format_line(f"   {COLOR_RED}•{COLOR_RESET} ... and {remaining} more warning(s)"))
    else:
        print(format_line(f" {COLOR_GREEN}✓ No warnings detected{COLOR_RESET}"))

    print(f"  {COLOR_BOLD}╚═══════════════════════════════════════════════════════════════════╝{COLOR_RESET}")
    print()
    draw_line()


# ======== REPORT GENERATION FUNCTIONS ========

def generate_node_report(nodes: List[Dict], cluster_info: Dict, output_file: str, quiet: bool = False):
    """Generate node hardware report CSV"""
    log_info(f"{COLOR_CYAN}Generating node hardware report...{COLOR_RESET}", quiet)

    csv_rows = []

    for node in nodes:
        # Basic node info
        node_ip = node.get('lanIP', '')
        backplane_ip = node.get('backplaneIP', '')

        # CPU info - cores and threads
        cpu_cores = node.get('numCores', 0)
        cpu_threads = node.get('numThreads', 0)
        cpu_info = f"{cpu_cores}/{cpu_threads}"

        # Memory in GB
        memory_bytes = node.get('memSize', 0)
        memory_gb = round(memory_bytes / 1_000_000_000, 2)

        # Storage capacity and usage from drives
        total_capacity = node.get('capacity', 0)
        capacity_tb = round(total_capacity / 1_000_000_000_000, 2)

        # Calculate used storage from drives
        used_bytes = 0
        drives = node.get('drives', [])
        for drive in drives:
            if drive.get('currentDisposition') == 'IN':
                used_bytes += drive.get('usedBytes', 0)
        used_tb = round(used_bytes / 1_000_000_000_000, 2)

        # CPU and Memory usage percentages
        cpu_usage = node.get('cpuUsage', 0)
        memory_usage = node.get('memUsagePercentage', 0)

        # Network status
        network_status = node.get('networkStatus', 'Unknown')

        # Disposition
        disposition = node.get('currentDisposition', 'Unknown')

        # ICOS Version
        icos_version = cluster_info.get('icosVersion', 'Unknown')

        csv_rows.append([
            node_ip,
            backplane_ip,
            cpu_info,
            memory_gb,
            capacity_tb,
            used_tb,
            f"{cpu_usage:.2f}" if cpu_usage else "0.00",
            f"{memory_usage:.2f}" if memory_usage else "0.00",
            network_status,
            disposition,
            icos_version
        ])

    # Write CSV file
    with open(output_file, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)

        writer.writerow([
            'Node IP', 'Backplane IP', 'CPU (cores/threads)', 'Memory (GB)',
            'Storage Capacity (TB)', 'Storage Used (TB)', 'CPU Usage (%)',
            'Memory Usage (%)', 'Network Status', 'Disposition', 'ICOS Version'
        ])

        writer.writerows(csv_rows)

    log_info(f"{COLOR_GREEN}✓ Node report saved to: {output_file}{COLOR_RESET}", quiet)


def generate_drive_report(nodes: List[Dict], output_file: str, quiet: bool = False):
    """Generate drive health report CSV"""
    log_info(f"{COLOR_CYAN}Generating drive health report...{COLOR_RESET}", quiet)

    csv_rows = []

    for node in nodes:
        node_ip = node.get('lanIP', '')
        drives = node.get('drives', [])

        for drive in drives:
            # Basic drive info
            slot = drive.get('slot', '')
            serial_number = drive.get('serialNumber', '')
            drive_type = drive.get('type', '')

            # Capacity and usage
            capacity_bytes = drive.get('capacityBytes', 0)
            capacity_gb = round(capacity_bytes / 1_000_000_000, 2)

            used_bytes = drive.get('usedBytes', 0)
            used_gb = round(used_bytes / 1_000_000_000, 2)

            # Usage percentage
            usage_pct = (used_bytes / capacity_bytes * 100) if capacity_bytes > 0 else 0

            # Temperature info
            temperature = drive.get('temperature', 0)
            max_temperature = drive.get('maxTemperature', 0)
            temperature_threshold = drive.get('temperatureThreshold', 0)

            # Health status
            is_healthy = drive.get('isHealthy', True)
            reallocated_sectors = drive.get('reallocatedSectors', 0)
            error_count = drive.get('errorCount', 0)

            # Determine health status
            if not is_healthy or error_count > 100:
                health_status = 'Failed'
            elif temperature > 0 and temperature_threshold > 0 and temperature > (temperature_threshold * 0.9):
                health_status = 'Warning'
            elif reallocated_sectors > 0:
                health_status = 'Warning'
            else:
                health_status = 'Healthy'

            # Disposition
            disposition = drive.get('currentDisposition', 'Unknown')

            csv_rows.append([
                node_ip,
                slot,
                serial_number,
                drive_type,
                capacity_gb,
                used_gb,
                f"{usage_pct:.2f}",
                temperature,
                max_temperature,
                temperature_threshold,
                health_status,
                reallocated_sectors,
                error_count,
                disposition
            ])

    # Write CSV file
    with open(output_file, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)

        writer.writerow([
            'Node IP', 'Slot', 'Serial Number', 'Type', 'Capacity (GB)',
            'Used (GB)', 'Usage (%)', 'Temperature (°C)', 'Max Temperature (°C)',
            'Temperature Threshold (°C)', 'Health Status', 'Reallocated Sectors',
            'Error Count', 'Disposition'
        ])

        writer.writerows(csv_rows)

    log_info(f"{COLOR_GREEN}✓ Drive report saved to: {output_file}{COLOR_RESET}", quiet)


def generate_iso_report(client: ScaleAPIClient, output_file: str, quiet: bool = False):
    """Generate ISO library inventory CSV"""
    log_info(f"{COLOR_CYAN}Generating ISO library report...{COLOR_RESET}", quiet)

    # Get ISO library data
    isos = client.get('ISO')
    if not isos:
        isos = []

    csv_rows = []

    for iso in isos:
        iso_name = iso.get('name', '')
        size_bytes = iso.get('size', 0)
        size_gb = round(size_bytes / 1_000_000_000, 2)
        path = iso.get('path', '')
        mounts = iso.get('mounts', 0)
        ready_for_insert = 'Yes' if iso.get('readyForInsert', False) else 'No'

        csv_rows.append([
            iso_name,
            size_gb,
            path,
            mounts,
            ready_for_insert
        ])

    # Sort by name
    csv_rows.sort(key=lambda x: x[0])

    # Write CSV file
    with open(output_file, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)

        writer.writerow([
            'ISO Name', 'Size (GB)', 'Path', 'Number of Mounts', 'Ready for Insert'
        ])

        writer.writerows(csv_rows)

    log_info(f"{COLOR_GREEN}✓ ISO report saved to: {output_file}{COLOR_RESET}", quiet)


# ======== EXCEL OUTPUT FUNCTIONS ========

def apply_excel_formatting(ws, header_row: int = 1, freeze_col: int = 2):
    """Apply standard Excel formatting to a worksheet"""
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Freeze panes (header row + first N columns)
    freeze_cell = f"{get_column_letter(freeze_col + 1)}{header_row + 1}"
    ws.freeze_panes = freeze_cell

    # Auto-filter on header row
    if ws.max_row > header_row:
        ws.auto_filter.ref = ws.dimensions

    # Auto-size columns (with max width limit)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        # Set width with min 10, max 50
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def add_conditional_formatting(ws, data_start_row: int = 2):
    """Add conditional formatting to worksheet based on data type"""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Color

    # Find State column (typically column E)
    state_col = None
    cpu_col = None

    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'State':
            state_col = idx
        elif cell.value == 'Node CPU Load (%)':
            cpu_col = idx

    # Format State column (Green for RUNNING, Red for STOPPED)
    if state_col:
        col_letter = get_column_letter(state_col)
        running_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        running_font = Font(color="006100")
        stopped_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        stopped_font = Font(color="9C0006")

        ws.conditional_formatting.add(
            f'{col_letter}{data_start_row}:{col_letter}{ws.max_row}',
            CellIsRule(operator='equal', formula=['"RUNNING"'], fill=running_fill, font=running_font)
        )
        ws.conditional_formatting.add(
            f'{col_letter}{data_start_row}:{col_letter}{ws.max_row}',
            CellIsRule(operator='notEqual', formula=['"RUNNING"'], fill=stopped_fill, font=stopped_font)
        )

    # Format CPU Load column (Red >80%, Yellow >60%)
    if cpu_col:
        col_letter = get_column_letter(cpu_col)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        ws.conditional_formatting.add(
            f'{col_letter}{data_start_row}:{col_letter}{ws.max_row}',
            CellIsRule(operator='greaterThan', formula=['80'], fill=red_fill)
        )
        ws.conditional_formatting.add(
            f'{col_letter}{data_start_row}:{col_letter}{ws.max_row}',
            CellIsRule(operator='between', formula=['60', '80'], fill=yellow_fill)
        )


def generate_summary_sheet(wb, cluster_info: Dict, nodes: List[Dict], stats: AuditStatistics, collection_time: str, warnings: List[str]):
    """Generate summary dashboard sheet"""
    ws = wb.create_sheet("Summary Dashboard", 0)  # Insert as first sheet

    # Set column width
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40

    row = 1

    # Title
    ws[f'A{row}'] = "SUPERAUDIT SUMMARY REPORT"
    ws[f'A{row}'].font = Font(size=16, bold=True, color="366092")
    row += 2

    # Collection Information
    ws[f'A{row}'] = "Collection Information"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    info_items = [
        ("Generated:", collection_time),
        ("Tool Version:", f"v{VERSION}"),
        ("Cluster Name:", cluster_info.get('name', 'Unknown')),
        ("Company:", cluster_info.get('company', 'Unknown')),
        ("ICOS Version:", cluster_info.get('icosVersion', 'Unknown')),
        ("Collection Time:", format_time(stats.get_elapsed_time()))
    ]

    for label, value in info_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    row += 1

    # Virtual Machine Statistics
    ws[f'A{row}'] = "Virtual Machine Statistics"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    running_pct = (stats.running_vms / stats.total_vms * 100) if stats.total_vms > 0 else 0
    stopped_pct = (stats.stopped_vms / stats.total_vms * 100) if stats.total_vms > 0 else 0

    vm_stats = [
        ("Total VMs:", f"{stats.total_vms}"),
        ("  Running:", f"{stats.running_vms} ({running_pct:.1f}%)"),
        ("  Stopped:", f"{stats.stopped_vms} ({stopped_pct:.1f}%)"),
        ("Memory Allocated:", f"{stats.total_memory_gb:,.1f} GB")
    ]

    for label, value in vm_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if not label.startswith('  '):
            ws[f'A{row}'].font = Font(bold=True)
        row += 1
    row += 1

    # Storage Statistics
    ws[f'A{row}'] = "Storage Statistics"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    storage_pct = stats.get_storage_usage_percent()
    storage_cap_tb = stats.total_storage_capacity_gb / 1000
    storage_used_tb = stats.total_storage_used_gb / 1000

    storage_stats = [
        ("Storage Allocated:", f"{storage_cap_tb:,.1f} TB"),
        ("Storage Used:", f"{storage_used_tb:,.1f} TB ({storage_pct:.1f}%)")
    ]

    for label, value in storage_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    row += 1

    # Snapshot & Backup Status
    ws[f'A{row}'] = "Snapshot & Backup Status"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    snap_pct = (stats.vms_with_snapshots / stats.total_vms * 100) if stats.total_vms > 0 else 0

    snapshot_stats = [
        ("VMs with Snapshots:", f"{stats.vms_with_snapshots} ({snap_pct:.1f}%)"),
        ("VMs without Snapshots:", f"{stats.vms_without_snapshots}")
    ]

    for label, value in snapshot_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    row += 1

    # Infrastructure
    ws[f'A{row}'] = "Infrastructure"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    total_drives = 0
    ssd_drives = 0
    hdd_drives = 0
    for node in nodes:
        for drive in node.get('drives', []):
            total_drives += 1
            if drive.get('type') == 'SSD':
                ssd_drives += 1
            else:
                hdd_drives += 1

    infra_stats = [
        ("Nodes:", f"{len(nodes)}"),
        ("Total Drives:", f"{total_drives} ({ssd_drives} SSD, {hdd_drives} HDD)")
    ]

    for label, value in infra_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    row += 1

    # Warnings
    if warnings:
        ws[f'A{row}'] = f"WARNINGS ({len(warnings)})"
        ws[f'A{row}'].font = Font(size=12, bold=True, color="9C0006")
        row += 1

        # Show first 20 warnings
        for warning in warnings[:20]:
            ws[f'A{row}'] = "⚠"
            ws[f'B{row}'] = warning
            ws[f'A{row}'].font = Font(color="9C0006")
            ws[f'B{row}'].font = Font(color="9C0006")
            row += 1

        if len(warnings) > 20:
            ws[f'B{row}'] = f"... and {len(warnings) - 20} more warning(s)"
            ws[f'B{row}'].font = Font(italic=True)
    else:
        ws[f'A{row}'] = "✓ No warnings detected"
        ws[f'A{row}'].font = Font(size=12, bold=True, color="006100")


def export_to_json(output_file: str, csv_rows: List, nodes: List[Dict], cluster_info: Dict, stats: AuditStatistics, quiet: bool = False):
    """Export VM data to JSON format (Push 4)"""
    import json

    log_info(f"{COLOR_CYAN}Exporting to JSON: {output_file}...{COLOR_RESET}", quiet)

    # Convert CSV rows to structured JSON
    headers = [
        'vm_uuid', 'mem_alloc_gb', 'type', 'cpus', 'state', 'node',
        'node_cpu_load_pct', 'name', 'description', 'tags',
        'boot_order', 'machine_type', 'operating_system', 'ha_policy',
        'drive_type', 'mounted_as', 'drive_size_gb', 'used_size_gb', 'ssd_tier',
        'cache_mode', 'disk_snapshots',
        'snaps', 'iso_mounted', 'vm_ip_address', 'vlans',
        'mac_addresses', 'adapter_types', 'connection_status',
        'snapshot_schedule_plan', 'replication_partner'
    ]

    vms = []
    for row in csv_rows:
        vm_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                vm_dict[header] = row[i]
        vms.append(vm_dict)

    # Build complete JSON structure
    data = {
        'cluster_info': cluster_info,
        'collection_time': datetime.datetime.now().isoformat(),
        'statistics': {
            'total_vms': stats.total_vms,
            'running_vms': stats.running_vms,
            'stopped_vms': stats.stopped_vms,
            'total_memory_gb': stats.total_memory_gb,
            'total_storage_capacity_gb': stats.total_storage_capacity_gb,
            'total_storage_used_gb': stats.total_storage_used_gb,
            'storage_usage_percent': stats.get_storage_usage_percent(),
            'vms_with_snapshots': stats.vms_with_snapshots,
            'vms_without_snapshots': stats.vms_without_snapshots,
            'vm_types': stats.vm_types,
            'avg_cpu_usage': stats.avg_cpu_usage,
            'max_cpu_usage': stats.max_cpu_usage,
            'avg_memory_usage': stats.avg_memory_usage,
            'max_memory_usage': stats.max_memory_usage
        },
        'vms': vms,
        'nodes': nodes
    }

    try:
        with open(output_file, 'w') as f:
            json.dump(data, f, indent=2)
        log_info(f"{COLOR_GREEN}✓ JSON export complete: {output_file}{COLOR_RESET}", quiet)
        return True
    except Exception as e:
        log_info(f"{COLOR_RED}Error exporting JSON: {e}{COLOR_RESET}", quiet)
        return False


def export_summary(output_file: str, cluster_info: Dict, stats: AuditStatistics, nodes: List[Dict], quiet: bool = False):
    """Export summary statistics to JSON (Push 4)"""
    import json

    log_info(f"{COLOR_CYAN}Exporting summary to: {output_file}...{COLOR_RESET}", quiet)

    summary = {
        'cluster': {
            'name': cluster_info.get('name', 'Unknown'),
            'company': cluster_info.get('company', 'Unknown'),
            'icos_version': cluster_info.get('icosVersion', 'Unknown'),
            'node_count': len(nodes)
        },
        'audit': {
            'timestamp': datetime.datetime.now().isoformat(),
            'total_vms': stats.total_vms,
            'running_vms': stats.running_vms,
            'stopped_vms': stats.stopped_vms
        },
        'resources': {
            'memory': {
                'total_allocated_gb': round(stats.total_memory_gb, 2)
            },
            'storage': {
                'total_capacity_gb': round(stats.total_storage_capacity_gb, 2),
                'total_used_gb': round(stats.total_storage_used_gb, 2),
                'total_free_gb': round(stats.get_storage_free_gb(), 2),
                'usage_percent': round(stats.get_storage_usage_percent(), 2),
                'free_percent': round(stats.get_storage_free_percent(), 2)
            }
        },
        'vm_types': stats.vm_types,
        'snapshots': {
            'vms_with_snapshots': stats.vms_with_snapshots,
            'vms_without_snapshots': stats.vms_without_snapshots,
            'snapshot_coverage_percent': round((stats.vms_with_snapshots / stats.total_vms * 100) if stats.total_vms > 0 else 0, 2)
        },
        'system_utilization': {
            'avg_cpu_usage_percent': round(stats.avg_cpu_usage, 2),
            'max_cpu_usage_percent': round(stats.max_cpu_usage, 2),
            'avg_memory_usage_percent': round(stats.avg_memory_usage, 2),
            'max_memory_usage_percent': round(stats.max_memory_usage, 2),
            'warnings': {
                'critical_count': len(stats.categorized_warnings['CRITICAL']),
                'warning_count': len(stats.categorized_warnings['WARNING']),
                'info_count': len(stats.categorized_warnings['INFO']),
                'total_count': stats.get_all_warnings_count()
            }
        }
    }

    try:
        with open(output_file, 'w') as f:
            json.dump(summary, f, indent=2)
        log_info(f"{COLOR_GREEN}✓ Summary export complete: {output_file}{COLOR_RESET}", quiet)
        return True
    except Exception as e:
        log_info(f"{COLOR_RED}Error exporting summary: {e}{COLOR_RESET}", quiet)
        return False


def generate_recommendation(warning: str, severity: str) -> str:
    """Generate recommendation based on warning (Push 3)"""
    warning_lower = warning.lower()

    # Snapshot recommendations
    if 'no snapshots' in warning_lower:
        return 'Configure snapshot schedule immediately. Snapshots are critical for data protection and recovery.'

    # HA policy recommendations
    if 'no ha policy' in warning_lower or 'ha policy' in warning_lower:
        return 'Configure High Availability policy to ensure VM redundancy across nodes.'

    # Replication recommendations
    if 'no replication' in warning_lower:
        return 'Consider configuring replication for disaster recovery and business continuity.'

    # Disk space recommendations
    if 'disk' in warning_lower and 'full' in warning_lower:
        if severity == 'CRITICAL':
            return 'URGENT: Free up disk space or expand storage immediately to prevent VM failures.'
        else:
            return 'Monitor disk usage closely and plan for storage expansion.'

    # CPU recommendations
    if 'cpu' in warning_lower:
        if severity == 'CRITICAL':
            return 'URGENT: High CPU usage may impact VM performance. Migrate VMs to other nodes or add capacity.'
        else:
            return 'Monitor CPU usage and consider load balancing VMs across nodes.'

    # Memory recommendations
    if 'memory' in warning_lower:
        return 'Monitor memory usage and consider adding memory to nodes or rebalancing VM placement.'

    # Generic recommendation
    return 'Review and address this issue to maintain optimal cluster health.'


def generate_excel_workbook(output_file: str, csv_rows: List, nodes: List[Dict], client: ScaleAPIClient,
                            cluster_info: Dict, stats: AuditStatistics, quiet: bool = False,
                            include_nodes: bool = True, include_drives: bool = True, include_isos: bool = True):
    """Generate Excel workbook with multiple sheets"""
    if not EXCEL_AVAILABLE:
        log_info(f"{COLOR_YELLOW}Warning: openpyxl not available. Install with: pip install openpyxl{COLOR_RESET}", quiet)
        log_info(f"{COLOR_YELLOW}Falling back to CSV output...{COLOR_RESET}", quiet)
        return False

    log_info(f"\n{COLOR_CYAN}Generating Excel workbook...{COLOR_RESET}", quiet)

    wb = Workbook()
    collection_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 1. VM Inventory Sheet
    log_info(f"{COLOR_CYAN}  - Creating VM Inventory sheet...{COLOR_RESET}", quiet)
    ws_vms = wb.create_sheet("VM Inventory")

    # Write headers (Enhanced with Push 2 fields)
    headers = [
        'VM UUID', 'Mem Alloc (GB)', 'Type', 'CPUs', 'State', 'Node',
        'Node CPU Load (%)', 'Name', 'Description', 'Tags',
        'Boot Order', 'Machine Type', 'Operating System', 'HA Policy',  # Enhanced VM fields
        'Drive Type', 'Mounted As', 'Drive Size (GB)', 'Used Size (GB)', 'SSD Tier',
        'Cache Mode', 'Disk Snapshots',  # Enhanced disk fields
        'Snaps', 'ISO Mounted', 'VM IP Address', 'VLANS',
        'MAC Addresses', 'Adapter Types', 'Connection Status',  # Enhanced network fields
        'Snapshot Schedule Plan', 'Replication Partner'
    ]
    ws_vms.append(headers)

    # Write data
    for row in csv_rows:
        ws_vms.append(row)

    # Apply formatting
    apply_excel_formatting(ws_vms, header_row=1, freeze_col=2)
    add_conditional_formatting(ws_vms, data_start_row=2)

    # 2. Node Hardware Sheet
    if include_nodes:
        log_info(f"{COLOR_CYAN}  - Creating Node Hardware sheet...{COLOR_RESET}", quiet)
        ws_nodes = wb.create_sheet("Node Hardware")

        # Headers
        node_headers = [
            'Node IP', 'Backplane IP', 'CPU (cores/threads)', 'Memory (GB)',
            'Storage Capacity (TB)', 'Storage Used (TB)', 'CPU Usage (%)',
            'Memory Usage (%)', 'Network Status', 'Disposition', 'ICOS Version'
        ]
        ws_nodes.append(node_headers)

        # Data
        for node in nodes:
            node_ip = node.get('lanIP', '')
            backplane_ip = node.get('backplaneIP', '')
            cpu_cores = node.get('numCores', 0)
            cpu_threads = node.get('numThreads', 0)
            cpu_info = f"{cpu_cores}/{cpu_threads}"
            memory_bytes = node.get('memSize', 0)
            memory_gb = round(memory_bytes / 1_000_000_000, 2)
            total_capacity = node.get('capacity', 0)
            capacity_tb = round(total_capacity / 1_000_000_000_000, 2)
            used_bytes = sum(drive.get('usedBytes', 0) for drive in node.get('drives', []) if drive.get('currentDisposition') == 'IN')
            used_tb = round(used_bytes / 1_000_000_000_000, 2)
            cpu_usage = node.get('cpuUsage', 0)
            memory_usage = node.get('memUsagePercentage', 0)
            network_status = node.get('networkStatus', 'Unknown')
            disposition = node.get('currentDisposition', 'Unknown')
            icos_version = cluster_info.get('icosVersion', 'Unknown')

            ws_nodes.append([
                node_ip, backplane_ip, cpu_info, memory_gb, capacity_tb, used_tb,
                round(cpu_usage, 2), round(memory_usage, 2), network_status, disposition, icos_version
            ])

        apply_excel_formatting(ws_nodes, header_row=1, freeze_col=1)

    # 3. Drive Health Sheet
    if include_drives:
        log_info(f"{COLOR_CYAN}  - Creating Drive Health sheet...{COLOR_RESET}", quiet)
        ws_drives = wb.create_sheet("Drive Health")

        # Headers
        drive_headers = [
            'Node IP', 'Slot', 'Serial Number', 'Type', 'Capacity (GB)',
            'Used (GB)', 'Usage (%)', 'Temperature (°C)', 'Max Temperature (°C)',
            'Temperature Threshold (°C)', 'Health Status', 'Reallocated Sectors',
            'Error Count', 'Disposition'
        ]
        ws_drives.append(drive_headers)

        # Data
        for node in nodes:
            node_ip = node.get('lanIP', '')
            for drive in node.get('drives', []):
                slot = drive.get('slot', '')
                serial_number = drive.get('serialNumber', '')
                drive_type = drive.get('type', '')
                capacity_bytes = drive.get('capacityBytes', 0)
                capacity_gb = round(capacity_bytes / 1_000_000_000, 2)
                used_bytes = drive.get('usedBytes', 0)
                used_gb = round(used_bytes / 1_000_000_000, 2)
                usage_pct = (used_bytes / capacity_bytes * 100) if capacity_bytes > 0 else 0
                temperature = drive.get('temperature', 0)
                max_temperature = drive.get('maxTemperature', 0)
                temperature_threshold = drive.get('temperatureThreshold', 0)
                is_healthy = drive.get('isHealthy', True)
                reallocated_sectors = drive.get('reallocatedSectors', 0)
                error_count = drive.get('errorCount', 0)

                # Determine health status
                if not is_healthy or error_count > 100:
                    health_status = 'Failed'
                elif temperature > 0 and temperature_threshold > 0 and temperature > (temperature_threshold * 0.9):
                    health_status = 'Warning'
                elif reallocated_sectors > 0:
                    health_status = 'Warning'
                else:
                    health_status = 'Healthy'

                disposition = drive.get('currentDisposition', 'Unknown')

                ws_drives.append([
                    node_ip, slot, serial_number, drive_type, capacity_gb, used_gb,
                    round(usage_pct, 2), temperature, max_temperature, temperature_threshold,
                    health_status, reallocated_sectors, error_count, disposition
                ])

        apply_excel_formatting(ws_drives, header_row=1, freeze_col=1)

    # 4. ISO Library Sheet
    if include_isos:
        log_info(f"{COLOR_CYAN}  - Creating ISO Library sheet...{COLOR_RESET}", quiet)
        ws_isos = wb.create_sheet("ISO Library")

        # Headers
        iso_headers = ['ISO Name', 'Size (GB)', 'Path', 'Number of Mounts', 'Ready for Insert']
        ws_isos.append(iso_headers)

        # Data
        try:
            isos = client.get('ISO')
            if isos:
                log_info(f"{COLOR_CYAN}    Found {len(isos)} ISO(s){COLOR_RESET}", quiet)
                iso_rows = []
                for iso in isos:
                    iso_name = iso.get('name', '')
                    size_bytes = iso.get('size', 0)
                    size_gb = round(size_bytes / 1_000_000_000, 2)
                    path = iso.get('path', '')
                    mounts = len(iso.get('mounts', []))  # Count the number of mounts
                    ready_for_insert = 'Yes' if iso.get('readyForInsert', False) else 'No'
                    iso_rows.append([iso_name, size_gb, path, mounts, ready_for_insert])

                # Sort by name
                iso_rows.sort(key=lambda x: x[0])

                for row in iso_rows:
                    ws_isos.append(row)
            else:
                log_info(f"{COLOR_YELLOW}    No ISOs found{COLOR_RESET}", quiet)
        except Exception as e:
            log_info(f"{COLOR_YELLOW}    Warning: Could not fetch ISO data: {e}{COLOR_RESET}", quiet)

        apply_excel_formatting(ws_isos, header_row=1, freeze_col=1)

    # ===== NEW SHEETS (Push 3) =====

    # 5. Warnings & Recommendations Sheet
    log_info(f"{COLOR_CYAN}  - Creating Warnings & Recommendations sheet...{COLOR_RESET}", quiet)
    ws_warnings = wb.create_sheet("Warnings & Recommendations")

    # Headers
    warning_headers = ['Severity', 'Issue', 'Recommendation']
    ws_warnings.append(warning_headers)

    # Data - categorized warnings with recommendations
    for severity in ['CRITICAL', 'WARNING', 'INFO']:
        for warning in stats.categorized_warnings[severity]:
            recommendation = generate_recommendation(warning, severity)
            ws_warnings.append([severity, warning, recommendation])

    apply_excel_formatting(ws_warnings, header_row=1, freeze_col=1)

    # Add conditional formatting for severity
    if stats.get_all_warnings_count() > 0:
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import CellIsRule

        # CRITICAL = red
        critical_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        ws_warnings.conditional_formatting.add(
            f'A2:A{len(stats.categorized_warnings["CRITICAL"]) + 1}',
            CellIsRule(operator='equal', formula=['"CRITICAL"'], fill=critical_fill)
        )

        # WARNING = yellow
        warning_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        ws_warnings.conditional_formatting.add(
            f'A2:A1000',
            CellIsRule(operator='equal', formula=['"WARNING"'], fill=warning_fill)
        )

        # INFO = light blue
        info_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        ws_warnings.conditional_formatting.add(
            f'A2:A1000',
            CellIsRule(operator='equal', formula=['"INFO"'], fill=info_fill)
        )

    # 6. Summary Dashboard (always first sheet)
    log_info(f"{COLOR_CYAN}  - Creating Summary Dashboard...{COLOR_RESET}", quiet)
    generate_summary_sheet(wb, cluster_info, nodes, stats, collection_time, stats.warnings)

    # Save workbook
    try:
        wb.save(output_file)
        log_info(f"{COLOR_GREEN}✓ Excel workbook saved to: {output_file}{COLOR_RESET}", quiet)
        return True
    except Exception as e:
        log_info(f"{COLOR_RED}Error saving Excel file: {e}{COLOR_RESET}", quiet)
        return False


# ======== MAIN PROGRAM ========

def determine_vm_type(vm: Dict) -> str:
    """
    Determine VM type based on VM properties and tags.

    Args:
        vm: VM dictionary from API

    Returns:
        VM type string (VM, VM REPLICA, TEMPLATE, PRODUCTION VM, DEV VM, TEST VM)
    """
    vm_name = vm.get('name', '')
    vm_tags = vm.get('tags', [])

    # Handle tags as list or string
    if isinstance(vm_tags, list):
        tags_str = '/'.join(vm_tags).lower()
    elif isinstance(vm_tags, str):
        tags_str = vm_tags.lower()
    else:
        tags_str = ''

    # Check if it's a replica
    if vm.get('replicationSourceVirDomainUUID'):
        return "VM REPLICA"

    # Check if it's a template (common naming conventions)
    elif 'template' in vm_name.lower() or 'tmpl' in vm_name.lower():
        return "TEMPLATE"

    # Check tags for special types
    elif tags_str:
        if 'template' in tags_str:
            return "TEMPLATE"
        elif 'test' in tags_str or 'testing' in tags_str:
            return "TEST VM"
        elif 'production' in tags_str or 'prod' in tags_str:
            return "PRODUCTION VM"
        elif 'dev' in tags_str or 'development' in tags_str:
            return "DEV VM"

    return "VM"  # Default


def extract_ha_policy(vm: Dict) -> str:
    """
    Extract HA policy from VM affinity strategy.

    Args:
        vm: VM dictionary from API

    Returns:
        HA policy string (STRICT, PREFERRED, or empty string)
    """
    ha_policy = vm.get('affinityStrategy', {})

    if isinstance(ha_policy, dict):
        ha_strategy = ha_policy.get('strictAffinity', '')
        ha_backup_node = ha_policy.get('preferredNodeUUID', '')

        # Simplified: STRICT, PREFERRED, or empty
        if ha_strategy == 'STRICT':
            return 'STRICT'
        elif ha_backup_node:
            return 'PREFERRED'

    return ''


def parse_arguments():
    """Parse command-line arguments"""
    parser = argparse.ArgumentParser(
        description='SuperAudit - Scale Computing HyperCore VM Inventory Tool',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument('output', nargs='?', default=DEFAULT_OUTPUT_FILE,
                        help='Output filename (default: superaudit_{clustername}.xlsx or .csv)')
    parser.add_argument('-n', '--node', help='Cluster node hostname or IP')
    parser.add_argument('-u', '--user', help='Username')
    parser.add_argument('-p', '--password', help='Password')
    parser.add_argument('-q', '--quiet', action='store_true',
                        help='Run in quiet mode (less verbose output)')
    parser.add_argument('-v', '--version', action='version',
                        version=f'SuperAudit v{VERSION}')
    parser.add_argument('--format', type=str, choices=['xlsx', 'csv'], default='xlsx',
                        help='Output format: xlsx (Excel) or csv (default: xlsx)')
    parser.add_argument('--no-verify-ssl', action='store_true',
                        help='Disable SSL certificate verification')
    parser.add_argument('--ca-bundle', type=str, metavar='PATH',
                        help='Path to CA certificate bundle for self-signed certificates')
    parser.add_argument('--include-nodes', action='store_true',
                        help='[Deprecated: auto-included in Excel] Generate node hardware report')
    parser.add_argument('--include-drives', action='store_true',
                        help='[Deprecated: auto-included in Excel] Generate drive health report')
    parser.add_argument('--include-isos', action='store_true',
                        help='[Deprecated: auto-included in Excel] Generate ISO library report')
    parser.add_argument('--all-reports', action='store_true',
                        help='[Deprecated: default in Excel] Generate all additional reports (CSV only)')
    parser.add_argument('--summary-only', action='store_true',
                        help='Show cluster and audit summary only (no file generation)')
    parser.add_argument('--warnings', action='store_true',
                        help='Show only warnings and recommendations (exit 0 if none, 1 if found)')

    # ===== FILTERING OPTIONS (Push 4) =====
    filter_group = parser.add_argument_group('Filtering Options (Push 4)')
    filter_group.add_argument('--filter-state', type=str, choices=['RUNNING', 'STOPPED', 'PAUSED'],
                              help='Filter VMs by state (RUNNING, STOPPED, PAUSED)')
    filter_group.add_argument('--filter-type', type=str,
                              help='Filter VMs by type (VM, "VM REPLICA", TEMPLATE, "PRODUCTION VM", "DEV VM", "TEST VM")')
    filter_group.add_argument('--filter-tag', type=str,
                              help='Filter VMs by tag (partial match)')
    filter_group.add_argument('--filter-name', type=str,
                              help='Filter VMs by name (partial match, case-insensitive)')
    filter_group.add_argument('--filter-node', type=str,
                              help='Filter VMs by node IP address')

    # ===== COMPARISON OPTIONS (Push 4) =====
    compare_group = parser.add_argument_group('Comparison Options (Push 4)')
    compare_group.add_argument('--compare', type=str, metavar='PREVIOUS_FILE',
                               help='Compare with previous audit file (Excel or CSV)')
    compare_group.add_argument('--compare-mode', type=str, choices=['changes', 'summary', 'full'], default='changes',
                               help='Comparison mode: changes (only differences), summary (stats only), full (complete)')

    # ===== EXPORT OPTIONS (Push 4) =====
    export_group = parser.add_argument_group('Export Options (Push 4)')
    export_group.add_argument('--export-json', type=str, metavar='FILE',
                              help='Export VM data to JSON file')
    export_group.add_argument('--export-summary', type=str, metavar='FILE',
                              help='Export summary statistics to JSON file')
    export_group.add_argument('--single-sheet', action='store_true',
                              help='CSV mode: Export single CSV instead of multiple files')

    # ===== PERFORMANCE & UX OPTIONS (Push 5) =====
    perf_group = parser.add_argument_group('Performance & UX Options (Push 5)')
    perf_group.add_argument('--dry-run', action='store_true',
                            help='Test run without generating files (shows what would be collected)')
    perf_group.add_argument('--test-connection', action='store_true',
                            help='Test API connection and exit (validates credentials and connectivity)')
    perf_group.add_argument('--log-file', type=str, metavar='FILE',
                            help='Write detailed log to file (includes debug information)')
    perf_group.add_argument('--config', type=str, metavar='FILE',
                            help='Load configuration from file (.superauditrc format)')
    perf_group.add_argument('--show-api-calls', action='store_true',
                            help='Show all API calls being made (debug mode)')

    # ===== DAEMON MODE OPTIONS (v8.0) =====
    daemon_group = parser.add_argument_group('Daemon Mode Options (v8.0 - Historical Data Collection)')
    daemon_group.add_argument('--daemon', action='store_true',
                              help='Run in daemon mode (continuous scheduled collection)')
    daemon_group.add_argument('--interval', type=int, default=15, metavar='MINUTES',
                              help='Collection interval in minutes for daemon mode (default: 15)')
    daemon_group.add_argument('--database', type=str, metavar='PATH',
                              help='Database path for historical data storage (required for --daemon)')
    daemon_group.add_argument('--log-to-db', action='store_true',
                              help='Log this audit run to database (one-time logging without daemon)')

    # ===== DASHBOARD MODE OPTIONS (v8.0) =====
    dashboard_group = parser.add_argument_group('Dashboard Mode Options (v8.0 - Web UI Visualization)')
    dashboard_group.add_argument('--dashboard', action='store_true',
                                 help='Start web dashboard server for historical data visualization')
    dashboard_group.add_argument('--dashboard-port', type=int, default=8080, metavar='PORT',
                                 help='Dashboard server port (default: 8080)')
    dashboard_group.add_argument('--dashboard-host', type=str, default='0.0.0.0', metavar='HOST',
                                 help='Dashboard server host (default: 0.0.0.0 - all interfaces)')

    return parser.parse_args()


def main():
    """Main program"""
    args = parse_arguments()

    # ===== DASHBOARD MODE ENTRY POINT (v8.0) =====
    if args.dashboard:
        # Import dashboard module
        try:
            from lib.dashboard.server import start_dashboard
        except ImportError as e:
            print(f"{COLOR_RED}Error: Missing required modules for dashboard mode:{COLOR_RESET}")
            print(f"{COLOR_RED}{e}{COLOR_RESET}")
            print(f"\n{COLOR_YELLOW}Install requirements: pip install -r requirements.txt{COLOR_RESET}\n")
            return 1

        # Validate database path
        if not args.database:
            print(f"{COLOR_RED}Error: --database is required for dashboard mode{COLOR_RESET}")
            print(f"{COLOR_YELLOW}Example: --dashboard --database /var/lib/superaudit/audit.db{COLOR_RESET}\n")
            return 1

        # Check if database exists
        from pathlib import Path
        if not Path(args.database).exists():
            print(f"{COLOR_RED}Error: Database file not found: {args.database}{COLOR_RESET}")
            print(f"{COLOR_YELLOW}Run daemon mode first to collect data, or use --log-to-db{COLOR_RESET}\n")
            return 1

        # Start dashboard server
        try:
            start_dashboard(
                database_path=args.database,
                host=args.dashboard_host,
                port=args.dashboard_port
            )
            return 0
        except KeyboardInterrupt:
            print("\nDashboard stopped")
            return 0
        except Exception as e:
            print(f"\n{COLOR_RED}Dashboard failed: {e}{COLOR_RESET}\n")
            return 1

    # ===== DAEMON MODE ENTRY POINT (v8.0) =====
    if args.daemon or args.log_to_db:
        # Import database and scheduler modules
        try:
            from lib.database import AuditDatabase
            from lib.scheduler import AuditDaemon
        except ImportError as e:
            print(f"{COLOR_RED}Error: Missing required modules for daemon mode:{COLOR_RESET}")
            print(f"{COLOR_RED}{e}{COLOR_RESET}")
            print(f"\n{COLOR_YELLOW}Install requirements: pip install -r requirements.txt{COLOR_RESET}\n")
            return 1

        # Validate database path
        if not args.database:
            print(f"{COLOR_RED}Error: --database is required for daemon mode or database logging{COLOR_RESET}")
            print(f"{COLOR_YELLOW}Example: --database /var/lib/superaudit/audit.db{COLOR_RESET}\n")
            return 1

        # Validate other required arguments
        if not args.node:
            print(f"{COLOR_RED}Error: --node is required for daemon mode{COLOR_RESET}")
            print(f"{COLOR_YELLOW}Example: --node 10.205.109.101{COLOR_RESET}\n")
            return 1

        # Create audit collection function for daemon
        def run_audit_and_log():
            """Run single audit and log to database"""
            # Run the normal audit collection (suppress file output)
            original_output = args.output
            original_quiet = args.quiet
            original_summary_only = args.summary_only

            # Override for daemon collection
            args.summary_only = True  # Don't generate files in daemon mode
            args.quiet = True  # Reduce verbosity

            try:
                # Call the normal main logic but capture the data
                # For now, we'll run a simplified version
                # TODO: Extract data collection into reusable function

                # Get credentials (same logic as main)
                username, password = None, None

                # Try command-line
                if args.user and args.password:
                    username, password = args.user, args.password

                # Try .netrc
                if not username or not password:
                    from pathlib import Path
                    netrc_creds = get_credentials_from_netrc(args.node)
                    if netrc_creds:
                        username, password = netrc_creds

                # Try environment
                if not username or not password:
                    env_creds = get_credentials_from_env()
                    if env_creds:
                        username, password = env_creds

                if not username or not password:
                    raise Exception("Credentials not found. Set SCALE_USER and SCALE_PASSWORD environment variables.")

                # Connect and collect data
                verify_ssl = not args.no_verify_ssl
                ca_bundle = args.ca_bundle

                client = ScaleAPIClient(args.node, username, password, verify_ssl, ca_bundle)
                client.login()

                # Collect cluster info
                cluster_info = get_cluster_info(client)
                nodes = get_nodes(client)
                vms = get_vms(client)
                conditions = get_conditions(client)
                replication_map = get_replication_connections(client)
                snapshot_schedules = get_snapshot_schedules(client)

                # Build node lookup
                node_uuid_to_lan = {node.get('uuid', ''): node.get('lanIP', '--') for node in nodes}

                # Initialize stats
                stats = AuditStatistics()

                # Collect node utilization metrics, storage from drives, and drive health
                for node in nodes:
                    cpu_usage = node.get('cpuUsage', 0.0)
                    memory_usage = node.get('memUsagePercentage', 0.0)
                    network_status = node.get('networkStatus', 'ONLINE')
                    drives = node.get('drives', [])
                    stats.add_node_utilization(cpu_usage, memory_usage, network_status, drives)

                    # Collect drive health for this node
                    for drive in drives:
                        is_healthy = drive.get('isHealthy', True)
                        stats.add_drive_health(is_healthy)

                # Process VMs (simplified for database logging)
                vm_list = []
                for vm in vms:
                    vm_uuid = vm.get('uuid', '')
                    vm_name = vm.get('name', 'Unknown')
                    vm_state = vm.get('state', 'UNKNOWN')
                    vm_mem = vm.get('mem', 0)
                    vm_cpu = vm.get('numVCPU', 0)

                    # Get VM type
                    vm_type = determine_vm_type(vm)

                    # Track stats
                    stats.add_vm(vm_state, vm_mem / (1024**3))
                    stats.add_vm_type(vm_type)

                    # Get node info
                    node_uuid = vm.get('nodeUUID', '')
                    node_lan_ip = node_uuid_to_lan.get(node_uuid, '--')

                    # Get node CPU
                    node_cpu_load = None
                    if vm_state == 'RUNNING' and node_uuid:
                        for node in nodes:
                            if node.get('uuid') == node_uuid:
                                node_cpu_load = node.get('cpuUsage')
                                break

                    # Snapshots
                    snap_uuids = vm.get('snapUUIDs', [])
                    has_snapshots = len(snap_uuids) > 0
                    stats.add_snapshot_status(has_snapshots)

                    # Replication
                    has_replication = vm_uuid in replication_map
                    replication_partners = replication_map.get(vm_uuid, '')

                    # Get disk info
                    block_devs = vm.get('blockDevs', [])
                    disk_total = 0
                    disk_used = 0
                    disk_snapshots = 0

                    for dev in block_devs:
                        if dev and dev.get('type') == 'VIRTIO_DISK':
                            capacity = dev.get('capacity', 0)
                            disk_total += capacity
                            disk_used += capacity  # Approximate

                            # Storage is now tracked from Node drives, not VM disks
                            # stats.add_disk(capacity / (1024**3), capacity / (1024**3))

                    # Network info
                    net_devs = vm.get('netDevs', [])
                    ip_addresses = []
                    vlans = []
                    mac_addresses = []

                    for net_dev in net_devs:
                        if not net_dev:
                            continue
                        ipv4_addrs = net_dev.get('ipv4Addresses', [])
                        ip_addresses.extend([ip for ip in ipv4_addrs if ip])

                        vlan = net_dev.get('vlan', 0)
                        if vlan:
                            vlans.append(str(vlan))

                        mac = net_dev.get('macAddress', '')
                        if mac:
                            mac_addresses.append(mac)

                    # Build VM dict for database
                    # Handle tags safely
                    tags_value = vm.get('tags', [])
                    if isinstance(tags_value, list):
                        tags_str = '/'.join(str(t) for t in tags_value if t)
                    else:
                        tags_str = str(tags_value) if tags_value else ''

                    # Handle boot devices safely
                    boot_devices = vm.get('bootDevices', [])
                    boot_order_str = '/'.join([
                        bd.get('type', '') if isinstance(bd, dict) else str(bd)
                        for bd in boot_devices if bd
                    ])

                    vm_data = {
                        'uuid': vm_uuid,
                        'name': vm_name,
                        'state': vm_state,
                        'vm_type': vm_type,
                        'description': vm.get('description', ''),
                        'tags': tags_str,
                        'cpu_count': vm_cpu,
                        'memory_bytes': vm_mem,
                        'disk_count': len(block_devs),
                        'disk_total_bytes': disk_total,
                        'disk_used_bytes': disk_used,
                        'disk_snapshot_bytes': disk_snapshots,
                        'boot_order': boot_order_str,
                        'machine_type': vm.get('machineType', ''),
                        'operating_system': vm.get('operatingSystem', ''),
                        'ha_policy': extract_ha_policy(vm),
                        'ip_addresses': ','.join(ip_addresses),
                        'vlans': ','.join(vlans),
                        'mac_addresses': ','.join(mac_addresses),
                        'network_adapter_count': len(net_devs),
                        'has_snapshots': has_snapshots,
                        'snapshot_count': len(snap_uuids),
                        'snapshot_schedules': snapshot_schedules.get(vm_uuid, '') if isinstance(snapshot_schedules, dict) else '',
                        'has_replication': has_replication,
                        'replication_partners': replication_partners if isinstance(replication_partners, str) else '',
                        'node_uuid': node_uuid,
                        'node_lan_ip': node_lan_ip,
                        'node_cpu_percent': node_cpu_load
                    }

                    vm_list.append(vm_data)

                # Build audit data structure
                audit_data = {
                    'cluster_info': {
                        'clusterName': cluster_info.get('name', 'Unknown'),
                        'clusterUUID': cluster_info.get('clusterUUID', ''),
                        'icosVersion': cluster_info.get('icosVersion', '')
                    },
                    'statistics': {
                        'nodes_count': len(nodes),
                        'nodes_online': stats.nodes_online,
                        'nodes_offline': stats.nodes_offline,
                        'vms_total': stats.total_vms,
                        'vms_running': stats.running_vms,
                        'vms_stopped': stats.stopped_vms,
                        'vms_paused': 0,  # Not tracked separately in AuditStatistics
                        'storage_allocated_bytes': int(stats.total_storage_capacity_gb * (1024**3)),
                        'storage_used_bytes': int(stats.total_storage_used_gb * (1024**3)),
                        'memory_allocated_bytes': int(stats.total_memory_gb * (1024**3)),
                        'avg_cpu_usage': stats.avg_cpu_usage,
                        'max_cpu_usage': stats.max_cpu_usage,
                        'avg_memory_usage': stats.avg_memory_usage,
                        'max_memory_usage': stats.max_memory_usage,
                        'total_drives': stats.total_drives,
                        'healthy_drives': stats.healthy_drives,
                        'unhealthy_drives': stats.unhealthy_drives,
                        'vms_with_snapshots': stats.vms_with_snapshots,
                        'warnings_critical': len(stats.categorized_warnings.get('CRITICAL', [])),
                        'warnings_warning': len(stats.categorized_warnings.get('WARNING', [])),
                        'warnings_info': len(stats.categorized_warnings.get('INFO', [])),
                        'warnings_total': stats.get_all_warnings_count(),
                        'execution_time': stats.get_elapsed_time()
                    },
                    'vms': vm_list,
                    'nodes': nodes,
                    'warnings': [],
                    'conditions': conditions
                }

                # Add warnings
                for severity, warning_list in stats.categorized_warnings.items():
                    for warning in warning_list:
                        audit_data['warnings'].append({
                            'severity': severity,
                            'message': warning
                        })

                # Store in database
                with AuditDatabase(args.database) as db:
                    snapshot_id = db.insert_audit_snapshot(audit_data)
                    print(f"✓ Stored audit snapshot ID: {snapshot_id} ({stats.total_vms} VMs, CPU: {stats.avg_cpu_usage:.1f}%, RAM: {stats.avg_memory_usage:.1f}%)")

                client.logout()
                return audit_data

            finally:
                # Restore original args
                args.output = original_output
                args.quiet = original_quiet
                args.summary_only = original_summary_only

        # Handle one-time database logging
        if args.log_to_db and not args.daemon:
            print(f"\n{COLOR_CYAN}Running audit with database logging...{COLOR_RESET}\n")
            try:
                run_audit_and_log()
                print(f"\n{COLOR_GREEN}✓ Audit logged to database: {args.database}{COLOR_RESET}\n")
                return 0
            except Exception as e:
                print(f"\n{COLOR_RED}✗ Failed to log audit: {e}{COLOR_RESET}\n")
                return 1

        # Handle daemon mode
        if args.daemon:
            daemon = AuditDaemon(
                audit_function=run_audit_and_log,
                interval_minutes=args.interval,
                database_path=args.database
            )

            try:
                daemon.start()
                return 0
            except KeyboardInterrupt:
                print("\nShutdown requested...")
                daemon.stop()
                return 0
            except Exception as e:
                print(f"\n{COLOR_RED}Daemon failed: {e}{COLOR_RESET}\n")
                return 1

    # ===== NORMAL ONE-TIME AUDIT MODE =====
    # Get connection details
    if args.node:
        host = args.node
    else:
        host = input("Cluster node hostname or IP: ")
        if not host:
            print("Error: Host is required")
            return 1

    # Secure credential gathering (in order of preference)
    username = None
    password = None

    # 1. Try command-line arguments (warn about security!)
    if args.user and args.password:
        username = args.user
        password = args.password
        if not args.quiet:
            print(f"{COLOR_YELLOW}⚠  WARNING: Passing passwords via -p is insecure!{COLOR_RESET}")
            print(f"{COLOR_YELLOW}   Password is visible in process list and shell history.{COLOR_RESET}")
            print(f"{COLOR_YELLOW}   Consider using environment variables or .netrc file instead.{COLOR_RESET}")
            print()

    # 2. Try .netrc file (secure, standard method)
    if not username or not password:
        netrc_creds = get_credentials_from_netrc(host)
        if netrc_creds:
            username, password = netrc_creds
            if not args.quiet:
                log_info(f"{COLOR_GREEN}✓ Using credentials from ~/.netrc{COLOR_RESET}", args.quiet)

    # 3. Try environment variables (secure for automation)
    if not username or not password:
        env_creds = get_credentials_from_env()
        if env_creds:
            username, password = env_creds
            if not args.quiet:
                log_info(f"{COLOR_GREEN}✓ Using credentials from environment variables{COLOR_RESET}", args.quiet)

    # 4. Fall back to command-line user only
    if not username and args.user:
        username = args.user

    # 5. Interactive prompts (most secure for human use)
    if not username:
        username = input("Username: ")
        if not username:
            print("Error: Username is required")
            return 1

    if not password:
        password = getpass("Password: ")
        if not password:
            print("Error: Password is required")
            return 1

    output_file = args.output
    quiet = args.quiet
    verify_ssl = not args.no_verify_ssl
    ca_bundle = args.ca_bundle

    # Validate ca_bundle if provided
    if ca_bundle:
        ca_bundle_path = Path(ca_bundle)
        if not ca_bundle_path.exists():
            print(f"{COLOR_RED}Error: CA bundle file not found: {ca_bundle}{COLOR_RESET}")
            return 1
        if not ca_bundle_path.is_file():
            print(f"{COLOR_RED}Error: CA bundle path is not a file: {ca_bundle}{COLOR_RESET}")
            return 1
        if not args.quiet:
            log_info(f"{COLOR_GREEN}✓ Using CA bundle: {ca_bundle}{COLOR_RESET}", quiet)

    # Special modes - force quiet if using --warnings
    summary_only = args.summary_only
    warnings_only = args.warnings
    if warnings_only:
        quiet = True  # Warnings mode is always quiet except for warnings

    # Print header
    if not quiet and not warnings_only:
        draw_line()
        print()
        draw_centered_text("S U P E R A U D I T")
        print()
        draw_centered_text("Scale Computing HyperCore VM Inventory Tool")
        print()
        draw_line()
        print()

    try:
        # Create API client and login
        log_info(f"{COLOR_CYAN}Connecting to {host}...{COLOR_RESET}", quiet)
        client = ScaleAPIClient(host, username, password, verify_ssl, ca_bundle)
        client.login()
        log_info(f"{COLOR_GREEN}✓ Logged in successfully{COLOR_RESET}", quiet)

        # ===== TEST CONNECTION MODE (Push 5) =====
        if args.test_connection:
            print(f"\n{COLOR_GREEN}{COLOR_BOLD}✓ CONNECTION TEST SUCCESSFUL{COLOR_RESET}\n")
            print(f"  Host:           {COLOR_YELLOW}{host}{COLOR_RESET}")
            print(f"  SSL Verify:     {COLOR_YELLOW}{'Enabled' if verify_ssl else 'Disabled'}{COLOR_RESET}")
            print(f"  CA Bundle:      {COLOR_YELLOW}{ca_bundle if ca_bundle else 'System default'}{COLOR_RESET}")
            print(f"  Username:       {COLOR_YELLOW}{username}{COLOR_RESET}")
            print(f"  Authentication: {COLOR_GREEN}Valid{COLOR_RESET}")

            # Try to get basic cluster info
            try:
                cluster_info = get_cluster_info(client)
                print(f"  Cluster Name:   {COLOR_YELLOW}{cluster_info.get('name', 'Unknown')}{COLOR_RESET}")
                print(f"  ICOS Version:   {COLOR_YELLOW}{cluster_info.get('icosVersion', 'Unknown')}{COLOR_RESET}")

                # Get quick stats
                nodes = get_nodes(client)
                vms = get_vms(client)
                print(f"  Nodes:          {COLOR_YELLOW}{len(nodes)}{COLOR_RESET}")
                print(f"  Virtual Machines: {COLOR_YELLOW}{len(vms)}{COLOR_RESET}")
            except Exception as e:
                print(f"  {COLOR_YELLOW}Note: Could not fetch cluster details: {e}{COLOR_RESET}")

            client.logout()
            print(f"\n{COLOR_GREEN}Test connection completed successfully.{COLOR_RESET}\n")
            return 0

        # Get cluster information
        log_info(f"{COLOR_CYAN}Getting cluster information...{COLOR_RESET}", quiet)
        cluster_info = get_cluster_info(client)

        # Determine output format
        output_format = args.format if EXCEL_AVAILABLE else 'csv'
        if output_format == 'xlsx' and not EXCEL_AVAILABLE:
            log_info(f"{COLOR_YELLOW}Warning: openpyxl not installed. Falling back to CSV format.{COLOR_RESET}", quiet)
            log_info(f"{COLOR_YELLOW}Install with: pip install openpyxl{COLOR_RESET}", quiet)
            output_format = 'csv'

        # Set default output filename based on cluster name and format if not provided
        if output_file is None:
            cluster_name = cluster_info.get('name', 'cluster')
            sanitized_name = sanitize_filename(cluster_name)
            extension = 'xlsx' if output_format == 'xlsx' else 'csv'
            output_file = f"superaudit_{sanitized_name}.{extension}"

        # If user specified output file without extension, add appropriate extension
        elif '.' not in output_file:
            extension = 'xlsx' if output_format == 'xlsx' else 'csv'
            output_file = f"{output_file}.{extension}"

        # Get nodes
        log_info(f"{COLOR_CYAN}Getting node information...{COLOR_RESET}", quiet)
        nodes = get_nodes(client)

        # Build node lookup maps
        node_uuid_to_lan = {}
        for node in nodes:
            node_uuid_to_lan[node.get('uuid', '')] = node.get('lanIP', '--')

        # Get VMs
        log_info(f"{COLOR_CYAN}Getting VM list...{COLOR_RESET}", quiet)
        vms = get_vms(client)
        num_vms = len(vms)

        # Get replication connections
        log_info(f"{COLOR_CYAN}Getting replication information...{COLOR_RESET}", quiet)
        replication_map = get_replication_connections(client)

        # Get snapshot schedules
        log_info(f"{COLOR_CYAN}Getting snapshot schedules...{COLOR_RESET}", quiet)
        snapshot_schedules = get_snapshot_schedules(client)

        # Initialize statistics tracker
        stats = AuditStatistics()

        # Collect node utilization metrics, storage from drives, and drive health
        for node in nodes:
            cpu_usage = node.get('cpuUsage', 0.0)
            memory_usage = node.get('memUsagePercentage', 0.0)
            network_status = node.get('networkStatus', 'ONLINE')
            drives = node.get('drives', [])
            stats.add_node_utilization(cpu_usage, memory_usage, network_status, drives)

            # Collect drive health for this node
            for drive in drives:
                is_healthy = drive.get('isHealthy', True)
                stats.add_drive_health(is_healthy)

        # Display system info and cluster summary
        if not quiet:
            print()
            print(f"  {COLOR_BOLD}System Information{COLOR_RESET}")
            draw_line()
            print(f"  Version: {COLOR_YELLOW}{VERSION}{COLOR_RESET}")
            print(f"  Cluster: {COLOR_YELLOW}{cluster_info['name']}{COLOR_RESET}")
            print(f"  Company: {COLOR_YELLOW}{cluster_info['company']}{COLOR_RESET}")
            draw_line()

        # Display cluster summary dashboard
        display_cluster_summary(cluster_info, nodes, quiet)

        # Task information
        if not quiet:
            print(f"  {COLOR_BOLD}Task Information{COLOR_RESET}")
            draw_line()
            print(f"  * Virtual Machines to Process: {COLOR_YELLOW}{num_vms}{COLOR_RESET}")
            print(f"  * Output File: {COLOR_YELLOW}{output_file}{COLOR_RESET}")
            draw_line()
            print()

        # Create CSV output
        csv_rows = []

        # Process each VM
        show_progress(0, num_vms, "", quiet, stats)

        for idx, vm in enumerate(vms, 1):
            vm_uuid = vm.get('uuid', '')
            vm_name = vm.get('name', 'Unknown')
            vm_description = vm.get('description', '')
            vm_mem = vm.get('mem', 0)
            vm_cpu = vm.get('numVCPU', 0)
            vm_state = vm.get('state', 'UNKNOWN')

            # Handle tags - could be a list or a string
            tags = vm.get('tags', [])
            if isinstance(tags, list):
                vm_tags = '/'.join(tags) if tags else ''
            elif isinstance(tags, str):
                vm_tags = tags
            else:
                vm_tags = ''

            # ===== ENHANCED VM DATA COLLECTION (Push 2) =====
            # Boot order - with defensive type checking
            boot_devices = vm.get('bootDevices', [])
            boot_order = '/'.join([bd.get('type', '') if isinstance(bd, dict) else str(bd) for bd in boot_devices if bd]) if boot_devices else ''

            # Machine type (BIOS or UEFI)
            machine_type = vm.get('machineType', '')

            # Operating system
            operating_system = vm.get('operatingSystem', '')

            # HA policy (affinity strategy)
            ha_policy = vm.get('affinityStrategy', {})
            if isinstance(ha_policy, dict):
                ha_strategy = ha_policy.get('strictAffinity', '')
                ha_backup_node = ha_policy.get('preferredNodeUUID', '')
                # Simplified: STRICT, PREFERRED, or empty
                if ha_strategy == 'STRICT':
                    vm_ha_policy = 'STRICT'
                elif ha_backup_node:
                    vm_ha_policy = 'PREFERRED'
                else:
                    vm_ha_policy = ''
            else:
                vm_ha_policy = ''

            # Get node info
            node_uuid = vm.get('nodeUUID', '')
            node_lan_ip = node_uuid_to_lan.get(node_uuid, '--')

            # Get node CPU load
            node_cpu_load = 'N/A'
            if vm_state == 'RUNNING' and node_uuid:
                # Find the node in our nodes list
                for node in nodes:
                    if node.get('uuid') == node_uuid:
                        cpu_usage = node.get('cpuUsage', None)
                        if cpu_usage is not None:
                            node_cpu_load = f"{cpu_usage:.2f}"
                        break

            # Get snapshot count from VM details
            # Note: Individual snapshot sizes are not available via REST API
            snap_uuids = vm.get('snapUUIDs', [])
            num_snaps = len(snap_uuids)

            # Get block devices
            block_devs = vm.get('blockDevs', [])

            # Get network devices and extract IPs/VLANs
            net_devs = vm.get('netDevs', [])
            ip_addresses = []
            vlans = []
            mac_addresses = []  # Enhanced: MAC addresses
            adapter_types = []  # Enhanced: Adapter types
            connection_statuses = []  # Enhanced: Connection status

            for net_dev in net_devs:
                if not net_dev:  # Skip null network devices
                    continue

                # Get IPv4 addresses
                ipv4_addrs = net_dev.get('ipv4Addresses', [])
                if ipv4_addrs:
                    for ipv4 in ipv4_addrs:
                        if ipv4:
                            ip_addresses.append(ipv4)

                # Get IPv6 addresses
                ipv6_addrs = net_dev.get('ipv6Addresses', [])
                if ipv6_addrs:
                    for ipv6 in ipv6_addrs:
                        if ipv6:
                            cleaned_ipv6 = clean_ipv6(ipv6)
                            if cleaned_ipv6:
                                ip_addresses.append(cleaned_ipv6)

                # Get VLAN - handle None and 0
                vlan = net_dev.get('vlan')
                if vlan is not None and vlan != 0:  # Don't include VLAN 0
                    vlans.append(str(vlan))

                # ===== ENHANCED NETWORK DATA (Push 2) =====
                # MAC address
                mac = net_dev.get('macAddress', '') or net_dev.get('mac', '')
                if mac:
                    mac_addresses.append(mac)

                # Adapter type (e.g., VIRTIO, RTL8139, E1000)
                adapter_type = net_dev.get('type', '')
                if adapter_type:
                    adapter_types.append(adapter_type)

                # Connection status
                connected = net_dev.get('connected', None)
                if connected is not None:
                    connection_statuses.append('Connected' if connected else 'Disconnected')

            vm_ip_addr = '/'.join(ip_addresses) if ip_addresses else ''
            vm_vlans = '/'.join(vlans) if vlans else ''
            vm_mac_addrs = '/'.join(mac_addresses) if mac_addresses else ''
            vm_adapter_types = '/'.join(adapter_types) if adapter_types else ''
            vm_conn_status = '/'.join(connection_statuses) if connection_statuses else ''

            # Get mounted ISOs
            iso_mounted = []
            for bd in block_devs:
                if not bd:  # Skip null block devices
                    continue
                bd_type = bd.get('type', '')
                if bd_type in ['IDE_CDROM', 'VIRTIO_CDROM']:
                    iso_name = bd.get('path', '')
                    if iso_name and iso_name.endswith('.iso'):
                        iso_mounted.append(iso_name.split('/')[-1])
            iso_mounted_str = '/'.join(iso_mounted) if iso_mounted else ''

            # Get snapshot schedule
            snapshot_schedule_uuid = vm.get('snapshotScheduleUUID', '')
            vm_snap_plan = snapshot_schedules.get(snapshot_schedule_uuid, '')

            # Get replication partner
            # Note: VMs may have remoteClusterConnectionUUID field linking to replication connection
            # Check for connection UUID first, fall back to direct lookup
            replication_conn_uuid = vm.get('remoteClusterConnectionUUID', '')
            replication_partner = replication_map.get(replication_conn_uuid, '')

            # If no connection found but VM has replication source, mark as replicated but unknown source
            if not replication_partner and vm.get('replicationSourceVirDomainUUID'):
                replication_partner = 'Replicated (source unknown)'

            # ===== ENHANCED VM TYPE DETECTION (Push 2) =====
            # More intelligent VM type categorization
            vm_type = "VM"  # Default

            # Check if it's a replica
            if vm.get('replicationSourceVirDomainUUID'):
                vm_type = "VM REPLICA"
            # Check if it's a template (common naming conventions)
            elif 'template' in vm_name.lower() or 'tmpl' in vm_name.lower():
                vm_type = "TEMPLATE"
            # Check tags for special types
            elif vm_tags:
                tags_lower = vm_tags.lower()
                if 'template' in tags_lower:
                    vm_type = "TEMPLATE"
                elif 'test' in tags_lower or 'testing' in tags_lower:
                    vm_type = "TEST VM"
                elif 'production' in tags_lower or 'prod' in tags_lower:
                    vm_type = "PRODUCTION VM"
                elif 'dev' in tags_lower or 'development' in tags_lower:
                    vm_type = "DEV VM"

            # ===== APPLY VM FILTERS (Push 4) =====
            if not vm_matches_filter(vm_name, vm_state, vm_type, vm_tags, node_lan_ip, args):
                continue  # Skip this VM if it doesn't match filter criteria

            # Track VM statistics (once per VM)
            mem_gb = float(convert_memory_to_gb(vm_mem))
            stats.add_vm(vm_state, mem_gb)
            stats.add_snapshot_status(num_snaps > 0)
            stats.add_vm_type(vm_type)  # Enhanced: Track VM types (Push 3)

            # ===== WARNING CHECKS =====

            # Check for missing snapshots (CRITICAL for running production VMs)
            if num_snaps == 0 and vm_state == 'RUNNING':
                severity = 'CRITICAL' if vm_type == 'PRODUCTION VM' else 'WARNING'
                stats.add_warning(f"{vm_name}: No snapshots configured (Running VM)", severity)

            # Check for missing HA policy (WARNING for production VMs)
            if not vm_ha_policy and vm_state == 'RUNNING' and vm_type == 'PRODUCTION VM':
                stats.add_warning(f"{vm_name}: No HA policy configured (Production VM)", 'WARNING')

            # Check for missing replication (INFO for production VMs)
            if not replication_partner and vm_type == 'PRODUCTION VM' and vm_state == 'RUNNING':
                stats.add_warning(f"{vm_name}: No replication configured (Production VM)", 'INFO')

            # Check node CPU (WARNING/CRITICAL based on threshold)
            if vm_state == 'RUNNING' and node_cpu_load != 'N/A':
                try:
                    cpu_val = float(node_cpu_load)
                    if cpu_val > 90:
                        stats.add_warning(f"Node {node_lan_ip}: Critical CPU usage ({cpu_val:.1f}%)", 'CRITICAL')
                    elif cpu_val > 80:
                        stats.add_warning(f"Node {node_lan_ip}: High CPU usage ({cpu_val:.1f}%)", 'WARNING')
                except:
                    pass

            # Process each block device (disk) - include all disk types
            disk_block_devs = [bd for bd in block_devs if 'DISK' in bd.get('type', '') and 'CDROM' not in bd.get('type', '')]

            if not disk_block_devs:
                # VM has no disks - still write a row
                csv_rows.append([
                    vm_uuid,
                    convert_memory_to_gb(vm_mem),
                    vm_type,
                    vm_cpu,
                    vm_state,
                    node_lan_ip,
                    node_cpu_load,
                    vm_name,
                    vm_description,
                    vm_tags,
                    boot_order,  # Enhanced: boot order
                    machine_type,  # Enhanced: machine type
                    operating_system,  # Enhanced: OS
                    vm_ha_policy,  # Enhanced: HA policy
                    '',  # drive type
                    '',  # mounted as
                    0,   # drive size
                    0,   # used size
                    '',  # ssd tier
                    '',  # cache mode (empty for no disk)
                    '',  # disk snapshots (empty for no disk)
                    num_snaps,
                    iso_mounted_str,
                    vm_ip_addr,
                    vm_vlans,
                    vm_mac_addrs,  # Enhanced: MAC addresses
                    vm_adapter_types,  # Enhanced: adapter types
                    vm_conn_status,  # Enhanced: connection status
                    vm_snap_plan,
                    replication_partner
                ])
            else:
                for bd in disk_block_devs:
                    drive_type = bd.get('type', '')
                    drive_slot = bd.get('slot', '')

                    # Map slot to mount point (simplified - may need adjustment)
                    if isinstance(drive_slot, str) and 'cdrom' in drive_slot.lower():
                        mount_point = 'no'
                    elif drive_slot:
                        mount_point = str(drive_slot)
                    else:
                        mount_point = 'no'

                    # Get VSD UUID
                    vsd_uuid = bd.get('uuid', '')

                    # Get disk details directly from block device
                    drive_size_gb = convert_to_gb(bd.get('capacity', 0))
                    used_capacity_gb = convert_to_gb(bd.get('allocation', 0))
                    ssd_priority = map_ssd_priority(bd.get('tieringPriorityFactor', 0))

                    # ===== ENHANCED DISK DATA (Push 2) =====
                    # Cache mode (writethrough, writeback, none)
                    cache_mode = bd.get('cacheMode', '')

                    # Disable snapshotting flag
                    disable_snaps = bd.get('disableSnapshotting', False)
                    disk_snapshots = 'Disabled' if disable_snaps else 'Enabled'

                    csv_rows.append([
                        vm_uuid,
                        convert_memory_to_gb(vm_mem),
                        vm_type,
                        vm_cpu,
                        vm_state,
                        node_lan_ip,
                        node_cpu_load,
                        vm_name,
                        vm_description,
                        vm_tags,
                        boot_order,  # Enhanced: boot order
                        machine_type,  # Enhanced: machine type
                        operating_system,  # Enhanced: OS
                        vm_ha_policy,  # Enhanced: HA policy
                        drive_type,
                        mount_point,
                        drive_size_gb,
                        used_capacity_gb,
                        ssd_priority,
                        cache_mode,  # Enhanced: cache mode
                        disk_snapshots,  # Enhanced: disk snapshot status
                        num_snaps,
                        iso_mounted_str,
                        vm_ip_addr,
                        vm_vlans,
                        vm_mac_addrs,  # Enhanced: MAC addresses
                        vm_adapter_types,  # Enhanced: adapter types
                        vm_conn_status,  # Enhanced: connection status
                        vm_snap_plan,
                        replication_partner
                    ])

                    # Storage is now tracked from Node drives, not VM disks
                    # stats.add_disk(drive_size_gb, used_capacity_gb)

                    # Check for full disks (Enhanced: with severity levels - Push 3)
                    if drive_size_gb > 0:
                        usage_pct = (used_capacity_gb / drive_size_gb) * 100
                        if usage_pct > 95:
                            stats.add_warning(f"{vm_name}: Disk {drive_slot} is {usage_pct:.0f}% full", 'CRITICAL')
                        elif usage_pct > 85:
                            stats.add_warning(f"{vm_name}: Disk {drive_slot} is {usage_pct:.0f}% full", 'WARNING')

            show_progress(idx, num_vms, vm_name, quiet, stats)

        # Clear progress bar
        if not quiet:
            print("\r" + " " * 100 + "\r", end='')

        # ===== DRY RUN MODE (Push 5) =====
        if args.dry_run:
            print(f"\n{COLOR_CYAN}{COLOR_BOLD}DRY RUN MODE - No files will be generated{COLOR_RESET}\n")
            print(f"  {COLOR_BOLD}Data Collection Summary:{COLOR_RESET}")
            print(f"  - VMs collected:     {COLOR_YELLOW}{len(csv_rows)}{COLOR_RESET} (from {num_vms} total)")
            print(f"  - Nodes scanned:     {COLOR_YELLOW}{len(nodes)}{COLOR_RESET}")
            print(f"  - VM types found:    {COLOR_YELLOW}{sum(1 for k, v in stats.vm_types.items() if v > 0)}{COLOR_RESET}")
            print(f"  - Warnings generated: {COLOR_YELLOW}{stats.get_all_warnings_count()}{COLOR_RESET}")
            print(f"\n  {COLOR_BOLD}Would generate files:{COLOR_RESET}")
            print(f"  - Main output:       {COLOR_YELLOW}{output_file}{COLOR_RESET}")
            if args.export_json:
                print(f"  - JSON export:       {COLOR_YELLOW}{args.export_json}{COLOR_RESET}")
            if args.export_summary:
                print(f"  - Summary export:    {COLOR_YELLOW}{args.export_summary}{COLOR_RESET}")
            print(f"\n  {COLOR_BOLD}Applied filters:{COLOR_RESET}")
            if args.filter_state or args.filter_type or args.filter_tag or args.filter_name or args.filter_node:
                if args.filter_state:
                    print(f"  - State:             {COLOR_YELLOW}{args.filter_state}{COLOR_RESET}")
                if args.filter_type:
                    print(f"  - Type:              {COLOR_YELLOW}{args.filter_type}{COLOR_RESET}")
                if args.filter_tag:
                    print(f"  - Tag:               {COLOR_YELLOW}{args.filter_tag}{COLOR_RESET}")
                if args.filter_name:
                    print(f"  - Name:              {COLOR_YELLOW}{args.filter_name}{COLOR_RESET}")
                if args.filter_node:
                    print(f"  - Node:              {COLOR_YELLOW}{args.filter_node}{COLOR_RESET}")
            else:
                print(f"  - {COLOR_GREEN}No filters applied (all VMs included){COLOR_RESET}")
            print(f"\n{COLOR_GREEN}Dry run completed successfully.{COLOR_RESET}\n")
            client.logout()
            return 0

        # Skip output generation if in summary-only or warnings-only mode
        if not summary_only and not warnings_only:
            # Choose output method based on format
            if output_format == 'xlsx':
                # Generate Excel workbook (all sheets included by default)
                success = generate_excel_workbook(
                    output_file, csv_rows, nodes, client, cluster_info, stats, quiet,
                    include_nodes=True,  # Always include in Excel
                    include_drives=True,  # Always include in Excel
                    include_isos=True  # Always include in Excel
                )

                if not success:
                    # Fall back to CSV if Excel generation failed
                    log_info(f"{COLOR_YELLOW}Falling back to CSV output...{COLOR_RESET}", quiet)
                    output_format = 'csv'
                    output_file = output_file.replace('.xlsx', '.csv')

            if output_format == 'csv':
                # Write CSV file (legacy format)
                log_info(f"\n{COLOR_CYAN}Writing CSV file...{COLOR_RESET}", quiet)

                with open(output_file, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)

                    # Write header
                    writer.writerow([
                        'VM UUID', 'Mem Alloc (GB)', 'Type', 'CPUs', 'State', 'Node',
                        'Node CPU Load (%)', 'Name', 'Description', 'Tags',
                        'Boot Order', 'Machine Type', 'Operating System', 'HA Policy',  # Enhanced fields
                        'Drive Type', 'Mounted As', 'Drive Size (GB)', 'Used Size (GB)', 'SSD Tier',
                        'Cache Mode', 'Disk Snapshots',  # Enhanced disk fields
                        'Snaps', 'ISO Mounted', 'VM IP Address', 'VLANS',
                        'MAC Addresses', 'Adapter Types', 'Connection Status',  # Enhanced network fields
                        'Snapshot Schedule Plan', 'Replication Partner'
                    ])

                    # Write data rows
                    writer.writerows(csv_rows)

                log_info(f"{COLOR_GREEN}✓ Main CSV file written successfully{COLOR_RESET}", quiet)

                # Generate additional CSV reports if requested
                if args.all_reports:
                    args.include_nodes = True
                    args.include_drives = True
                    args.include_isos = True

                # Get sanitized cluster name for fallback filenames
                cluster_name = cluster_info.get('name', 'cluster')
                sanitized_name = sanitize_filename(cluster_name)

                if args.include_nodes:
                    node_output = output_file.replace('.csv', '_nodes.csv')
                    if not node_output.endswith('_nodes.csv'):
                        node_output = f'superaudit_{sanitized_name}_nodes.csv'
                    generate_node_report(nodes, cluster_info, node_output, quiet)

                if args.include_drives:
                    drive_output = output_file.replace('.csv', '_drives.csv')
                    if not drive_output.endswith('_drives.csv'):
                        drive_output = f'superaudit_{sanitized_name}_drives.csv'
                    generate_drive_report(nodes, drive_output, quiet)

                if args.include_isos:
                    iso_output = output_file.replace('.csv', '_isos.csv')
                    if not iso_output.endswith('_isos.csv'):
                        iso_output = f'superaudit_{sanitized_name}_isos.csv'
                    generate_iso_report(client, iso_output, quiet)

        # ===== ADDITIONAL EXPORT OPTIONS (Push 4) =====

        # JSON export
        if args.export_json:
            export_to_json(args.export_json, csv_rows, nodes, cluster_info, stats, quiet)

        # Summary export
        if args.export_summary:
            export_summary(args.export_summary, cluster_info, stats, nodes, quiet)

        # Logout
        client.logout()

        # Handle special display modes
        if warnings_only:
            # Warnings-only mode: Show only warnings
            if stats.warnings:
                print(f"\n{COLOR_RED}{COLOR_BOLD}⚠  WARNINGS DETECTED:{COLOR_RESET}")
                for idx, warning in enumerate(stats.warnings, 1):
                    print(f"  {COLOR_RED}{idx}.{COLOR_RESET} {warning}")
                print(f"\n{COLOR_YELLOW}Total warnings: {len(stats.warnings)}{COLOR_RESET}\n")
                return 1  # Exit code 1 when warnings are found
            else:
                print(f"{COLOR_GREEN}✓ No warnings detected{COLOR_RESET}")
                return 0  # Exit code 0 when no warnings
        else:
            # Normal or summary-only mode: Display full summary
            display_summary_report(stats, output_file, quiet)
            return 0

    except Exception as e:
        print(f"\n{COLOR_RED}Error: {e}{COLOR_RESET}", file=sys.stderr)
        return 1


if __name__ == '__main__':
    sys.exit(main())
